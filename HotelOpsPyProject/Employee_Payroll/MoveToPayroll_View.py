from app.Global_Api import get_department_bulk_names, get_designation_bulk_names
from Employee_Payroll.models import Raw_Attendance_Data, Attendance_Data
# from app.views import EmployeeDataSelect,EmployeeDataSelectForSalary
# from django.shortcuts import redirect, get_object_or_404
from app.models import OrganizationMaster,EmployeeMaster
# from django.template.loader import get_template
from datetime import date, timedelta, datetime
from django.shortcuts import render,redirect
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from calendar import month_name
from .models import * 
import json
import uuid
from django.db.models import Q
import calendar
from rest_framework.views import APIView



from django.db  import connection, transaction

def Employee_Data_Select_Payroll(OrganizationID=None, EmployeeCode=None,Designation =None,ReportingtoDesignation =None):
    with connection.cursor() as cursor:
        cursor.execute("EXEC SP_EmployeeMaster_For_Leave_Api @OrganizationID=%s, @EmployeeCode=%s, @Designation=%s, @ReportingtoDesignation=%s", [OrganizationID, EmployeeCode,Designation,ReportingtoDesignation])
        rows = cursor.fetchall()
        columns = [col[0] for col in cursor.description]
    rowslist = [dict(zip(columns, row)) for row in rows]
    return rowslist



class Payroll_Employee_Data_API(APIView):
    def get(self, request):
        try:
            OrganizationID = request.query_params.get('OID')
            EmployeeCode = request.query_params.get('EmployeeCode')
            Designation = request.query_params.get('Designation')
            ReportingtoDesignation = request.query_params.get('ReportingtoDesignation')
            
            # print('OrganizationID:',OrganizationID)

            data = Employee_Data_Select_Payroll(
                OrganizationID=OrganizationID,
                EmployeeCode=EmployeeCode,
                Designation=Designation,
                ReportingtoDesignation=ReportingtoDesignation
            )

            return Response({
                "success": True,
                "count": len(data),
                "data": data
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "success": False,
                "message": str(e)
            }, status=status.HTTP_400_BAD_REQUEST)


# ------------- View Function
from decimal import Decimal, ROUND_HALF_UP

def round_decimal(value):
    return value.quantize(Decimal('1'), rounding=ROUND_HALF_UP)

def round__Value_decimal(val):
    return Decimal(str(val)).quantize(Decimal("0.00"), rounding=ROUND_HALF_UP)

def MoveToPayroll_View(request):
    OrganizationID = request.GET.get("Organizations") 
    S_OID = request.session.get("OrganizationID")
    Departments = request.GET.get("Departments") 
    Departments_list = Departments.split(",") if Departments else []
    Designations = request.GET.get("Designations") 
    Designations_list = Designations.split(",") if Designations else []

    Status = request.GET.get("Status", 'all')
    Status_list = Status.split(",") if Status else []

    FromDate = request.GET.get("FromDate")
    ToDate = request.GET.get("ToDate")

    if OrganizationID is None or not OrganizationID:
        OrganizationID = S_OID

    department_names = get_department_bulk_names(Departments_list) if Departments_list else []
    designation_names = get_designation_bulk_names(Designations_list) if Designations_list else []

    
    # EmployeesList = EmployeeMaster.objects.filter(
    #     OrganizationID=OrganizationID,
    #     IsDelete=False, 
    #     IsSecondary=False,
    #     EmpStatus__in=["On Probation", "Confirmed", "Not Confirmed","Resigned","F&F In process","Absconding", "Terminate"]
    # )

    if Status_list is not None:
        if "all" in Status_list:
            # print("there is any all status is here")
            pass
            # don't filter, show all
        else:
            # print("there is any other status")
            EmployeesList = EmployeesList.filter(EmpStatus__in=Status_list)


    if department_names:
        EmployeesList = EmployeesList.filter(Department__in=department_names)

    if designation_names:
        EmployeesList = EmployeesList.filter(Designation__in=designation_names)

    # EmployeesList = EmployeesList.order_by("EmployeeCode", "EmpName").values(
    #     'id', 'EmpName', 'EmployeeCode', 'Department', 'Designation','EmpStatus','DateofJoining','Level','EmpID'
    # )
    
    EmployeesList = Employee_Data_Select_Payroll(
        OrganizationID=OrganizationID,
        EmployeeCode=None,
        Designation=Designations,
        ReportingtoDesignation=None
    )

    now = timezone.now()
    today_date_obj = now.date()
    # this_month_obj = now.month()
    # today_date_obj = date(2025, 8, 31)
    today_Month_Name = month_name[today_date_obj.month].lower().capitalize()

    current_year = datetime.now().year
    years = list(range(2024, current_year + 2))

    # Get the first day of the current month
    month_start = date(now.year, now.month, 1)
    last_day = calendar.monthrange(now.year, now.month)[1]
    month_end = date(now.year, now.month, last_day)

    # print("this month month_start is here:", month_start)
    # print("this month month_end is here:", month_end)

    # print("today is here:", today_date_obj)
    # print("Session Organization ID :", OrganizationID)
    context = {
        "EmployeesList": EmployeesList,
        "Filter_Departments": json.dumps(Departments_list),    
        "Filter_Designations": json.dumps(Designations_list),  
        "Status_list": json.dumps(Status_list), 
        "FromDate": FromDate,
        "ToDate": ToDate,
        'Session_OrganizationID': S_OID,
        'Selected_OID': OrganizationID,
        # 'today': today,
        'today': today_date_obj,
        'today_Month_Name': today_Month_Name,
        'years': years, 
        'current_year': current_year,
        'month_end': month_end,
        'month_start': month_start,
    }
    return render(request, "EMP_PAY/MoveToPayroll_Template/MoveToPayroll.html", context)



# Helper Function
def get_present_value(status: str) -> float:
    """
    Return PresentValue based on attendance status.
        1.0  → Full Present
        0.5  → Half Day Present
        0.0  → Absent or LWP
    """
    non_present_statuses = ["Absent", "LWP"]

    if status == "Half Day Present":
        return 0.5
    elif status not in non_present_statuses:
        return 1.0
    return 0.0


# --------------- Attedence Transfering 
def start_move_to_payroll(request):
    if request.method == "POST":
        OrganizationID = request.POST.get("SelectedOID")
        selected_employee_codes = request.POST.getlist("all_emp_codes")
        print("selected employee here:", selected_employee_codes)
        # selected_employee_ids = request.POST.getlist("all_emp_ids")
        FromDate = request.POST.get("FromDate")
        ToDate = request.POST.get("ToDate")
        Exisiting_Data_check = request.POST.get("Exisiting_Data_check")
        UserID = str(request.session["UserID"]) or 0
        S_OID = request.session.get("OrganizationID") or 0
        SelectMonth = request.POST.get("SelectMonth")
        SelectYear = request.POST.get("SelectYear")

        employee_data = json.loads(request.POST.get("employee_data", "[]"))

        selected_employee_codes = [emp["emp_code"] for emp in employee_data]
        selected_employee_ids = [emp["emp_id"] for emp in employee_data]

        if not FromDate or not ToDate:
            return JsonResponse({"error": "FromDate and ToDate are required"}, status=400)

        from_date = timezone.datetime.strptime(FromDate, "%Y-%m-%d").date()
        to_date = timezone.datetime.strptime(ToDate, "%Y-%m-%d").date()
        total_days = (to_date - from_date).days + 1
        total_entry = len(selected_employee_codes) * total_days

        task_id = str(uuid.uuid4())  # unique task
        progress = PayrollProgress.objects.create(
            task_id=task_id, 
            total=total_entry, 
            processed=0, 
            inserted=0, 
            status="running"
        )

        import threading
        t = threading.Thread(
            target=process_payroll_task,
            args=(task_id, OrganizationID, selected_employee_codes, selected_employee_ids, from_date, to_date, SelectMonth, SelectYear, Exisiting_Data_check, UserID, S_OID),
        )
        t.start()

        return JsonResponse({"task_id": task_id})



def process_payroll_task(task_id, OrganizationID, selected_employee_codes, selected_employee_ids, from_date, to_date, SelectMonth, SelectYear, Exisiting_Data_check, UserID, S_OID):

    PayrollProgress.objects.get(task_id=task_id)

    inserted_count = 0
    process_entry = 0

    current_date = from_date
    while current_date <= to_date:

        # monthno = current_date.month
        # year = current_date.year
        # monthname = month_name[monthno]  # "August", "September", etc.

        monthname = SelectMonth
        year = SelectYear
        # monthno = month_name[monthno]  # "August", "September", etc.
        month_number = datetime.strptime(monthname, "%B").month
        monthno = month_number
        # print("monthno is here:", monthno)
        
        for emp_code in selected_employee_codes:
            process_entry += 1

            attendance = Attendance_Data.objects.filter(
                IsDelete=False,
                OrganizationID=OrganizationID,
                EmployeeCode=emp_code,
                Date=current_date
            ).first()

            salary_exists = SalaryAttendance.objects.filter(
                OrganizationID=OrganizationID,
                EmployeeCode=emp_code,
                Date=current_date
            ).exists()

            if attendance is None:
                if not salary_exists or Exisiting_Data_check:
                    if salary_exists and Exisiting_Data_check:
                        SalaryAttendance.objects.filter(
                            OrganizationID=OrganizationID,
                            EmployeeCode=emp_code,
                            Date=current_date
                        ).delete()

                    SalaryAttendance.objects.create(
                        # EmpID=empid_map.get(emp_code, 0),
                        EmpID=selected_employee_ids[selected_employee_codes.index(emp_code)] if emp_code in selected_employee_codes else 0,
                        EmployeeCode=emp_code,
                        Date=current_date,
                        Status="Absent",
                        ActualStatus="Absent",
                        PresentValue=0.0,   
                        IsPresent=False, 
                        From_Date=from_date,
                        To_Date=to_date,
                        OrganizationID=S_OID,
                        LeaveID=0,
                        HotelID=OrganizationID,
                        CreatedBy=UserID,
                        CreatedDateTime=timezone.now(),
                        month=monthno,
                        year=year,
                        month_name=monthname,
                        IsAttendanceMoved = True,
                    )
                    inserted_count += 1

            else:
                if not salary_exists or Exisiting_Data_check:
                    if salary_exists and Exisiting_Data_check:
                        SalaryAttendance.objects.filter(
                            OrganizationID=OrganizationID,
                            EmployeeCode=emp_code,
                            Date=current_date
                        ).delete()

                    SalaryAttendance.objects.create(
                        # EmpID=empid_map.get(emp_code, 0),
                        EmpID=selected_employee_ids[selected_employee_codes.index(emp_code)] if emp_code in selected_employee_codes else 0,
                        EmployeeCode=attendance.EmployeeCode,
                        Date=attendance.Date,
                        In_Time=attendance.In_Time,
                        Out_Time=attendance.Out_Time,
                        S_In_Time=attendance.S_In_Time,
                        S_Out_Time=attendance.S_Out_Time,
                        Duty_Hour=attendance.Duty_Hour,
                        Status=attendance.Status,
                        ActualStatus=attendance.Status,
                        PresentValue=get_present_value(attendance.Status),   
                        IsPresent=get_present_value(attendance.Status) > 0,  
                        LeaveID=0,
                        From_Date=from_date,
                        To_Date=to_date,
                        OrganizationID=S_OID,
                        HotelID=attendance.OrganizationID,
                        CreatedBy=UserID,
                        CreatedDateTime=timezone.now(),
                        IsAttendanceMoved = True,

                        month=monthno,
                        year=year,
                        month_name=monthname,
                    )
                    inserted_count += 1

            if process_entry % 10 == 0:
                PayrollProgress.objects.filter(task_id=task_id).update(
                    processed=process_entry,
                    inserted=inserted_count
                )

        current_date += timedelta(days=1)

    PayrollProgress.objects.filter(task_id=task_id).update(
        processed=process_entry,
        inserted=inserted_count,
        status="done"
    )


from django.http import JsonResponse
from django.shortcuts import get_object_or_404

def payroll_progress(request, task_id):
    # print("I am called this api--->")
    try:
        progress = PayrollProgress.objects.get(task_id=task_id)
        # print("processed::", progress.processed)
        # print("total::", progress.total)
        # print("inserted::", progress.inserted)
        # print("status::", progress.status)
        return JsonResponse({
            "processed": progress.processed,
            "total": progress.total,
            "inserted": progress.inserted,
            "status": progress.status
        })
    except Exception as e:
        # log the exact error for debugging
        # print("Error in payroll_progress:", str(e))
        return JsonResponse({"error": str(e)}, status=500)









#  ---------- Attendance  Lock


# def Attendance_Lock_Payroll_View(request):
#     OrganizationID = request.GET.get("Organizations") 
#     S_OID = request.session.get("OrganizationID")
#     Departments = request.GET.get("Departments") 
#     Departments_list = Departments.split(",") if Departments else []
#     Designations = request.GET.get("Designations") 
#     Designations_list = Designations.split(",") if Designations else []

#     Status = request.GET.get("Status", 'all')
#     Status_list = Status.split(",") if Status else []

#     FromDate = request.GET.get("FromDate")
#     ToDate = request.GET.get("ToDate")

#     if OrganizationID is None or not OrganizationID:
#         OrganizationID = S_OID

#     department_names = get_department_bulk_names(Departments_list) if Departments_list else []
#     designation_names = get_designation_bulk_names(Designations_list) if Designations_list else []

    
#     EmployeesList = EmployeeMaster.objects.filter(
#         OrganizationID=OrganizationID,
#         IsDelete=False, 
#         IsSecondary=False,
#         EmpStatus__in=["On Probation", "Confirmed", "Not Confirmed"]
#     )


#     if Status_list is not None:
#         if "all" in Status_list:
#             pass
#         else:
#             EmployeesList = EmployeesList.filter(EmpStatus__in=Status_list)


#     if department_names:
#         EmployeesList = EmployeesList.filter(Department__in=department_names)

#     if designation_names:
#         EmployeesList = EmployeesList.filter(Designation__in=designation_names)

#     EmployeesList = EmployeesList.order_by("EmployeeCode", "EmpName").values(
#         'id', 'EmpName', 'EmployeeCode', 'Department', 'Designation','EmpStatus','DateofJoining','Level'
#     )


#     now = timezone.now()
#     today_date_obj = now.date()


#     # print("today is here:", today_date_obj)
#     # print("Session Organization ID :", OrganizationID)
#     context = {
#         "EmployeesList": EmployeesList,
#         "Filter_Departments": json.dumps(Departments_list),    
#         "Filter_Designations": json.dumps(Designations_list),  
#         "Status_list": json.dumps(Status_list), 
#         "FromDate": FromDate,
#         "ToDate": ToDate,
#         'Session_OrganizationID': S_OID,
#         'Selected_OID': OrganizationID,
#         'today': today_date_obj,
#     }
#     return render(request, "EMP_PAY/MoveToPayroll_Template/Attendance_Lock_Payroll.html", context)

from django.urls import reverse
from django.shortcuts import redirect

def Attendance_Lock_Payroll_View(request):
    if 'OrganizationID' not in request.session:
        return redirect(MasterAttribute.Host)
    else:
        print("Show Page Session")

    OrganizationID = request.session["OrganizationID"]
    # UserID = str(request.session["UserID"])
    # UserType = request.session["UserType"]
    current_date = datetime.now()


    year = request.GET.get('year')
    month_no = request.GET.get('month_no')

    if not year or not month_no:
        url = reverse("Attendance_Lock_Payroll")  
        return redirect(f"{url}?year={current_date.year}&month_no={current_date.month}")
    

    year = int(year)
    month_no = int(month_no)

    previous_month = datetime(year, month_no, 1) - timedelta(days=1)
    next_month = (datetime(year, month_no, 28) + timedelta(days=4)).replace(day=1)

    # Prevent future months
    if next_month > datetime.now().replace(day=1):
        next_month = datetime.now().replace(day=1)

    context = {
        'current_month': datetime(year, month_no, 1),
        'previous_month': previous_month,
        'next_month': next_month,
        'month_no': month_no,
        'year': year,
        'Session_OrganizationID':OrganizationID
    }
    return render(request, "EMP_PAY/MoveToPayroll_Template/Attendance_Lock_Payroll.html", context)


def Generte_Salary_View(request):
    if 'OrganizationID' not in request.session:
        return redirect(MasterAttribute.Host)
    else:
        print("Show Page Session")

    OrganizationID = request.session["OrganizationID"]
    # UserID = str(request.session["UserID"])
    # UserType = request.session["UserType"]
    current_date = datetime.now()


    year = request.GET.get('year')
    month_no = request.GET.get('month_no')

    if not year or not month_no:
        url = reverse("Generte_Salary_View")  
        return redirect(f"{url}?year={current_date.year}&month_no={current_date.month}")
    

    year = int(year)
    month_no = int(month_no)

    previous_month = datetime(year, month_no, 1) - timedelta(days=1)
    next_month = (datetime(year, month_no, 28) + timedelta(days=4)).replace(day=1)

    # Prevent future months
    if next_month > datetime.now().replace(day=1):
        next_month = datetime.now().replace(day=1)

    print("hotel id is here:", OrganizationID)

    context = {
        'current_month': datetime(year, month_no, 1),
        'previous_month': previous_month,
        'next_month': next_month,
        'month_no': month_no,
        'year': year,
        'Session_OrganizationID':OrganizationID
    }
    return render(request, "EMP_PAY/MoveToPayroll_Template/Generate_Salary_Temp.html", context)





from app.views import EmployeeDataSelect,EmployeeDataSelectForSalary
import calendar
from Leave_Management_System.models import  Leave_Type_Master,Emp_Leave_Balance_Master

from collections import defaultdict
    
from .models import SalarySlip
from collections import defaultdict
from datetime import datetime, timedelta
from django.shortcuts import render, redirect

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import get_object_or_404
from datetime import datetime, timedelta
import calendar
from itertools import groupby
from operator import attrgetter


# -----  get_month_range

def get_month_range(year, month_no):
    start_date = datetime(year, month_no, 1)
    next_month = start_date.replace(day=28) + timedelta(days=4)
    end_date = next_month - timedelta(days=next_month.day)
    return start_date, end_date

# ----- get_previous_month_start_date ----

def get_previous_month_start_date(year, month_no):
    if month_no == 1:
        year -= 1
        month_no = 12
    else:
        month_no -= 1
    return datetime(year, month_no, 26)

# ---- get_attendance_counts

def get_attendance_counts(attendance, leavelist):
    week_off_count = 0
    present_count = 0
    absent_count = 0
    leave_counts = {leave: 0 for leave in leavelist}

    for att in attendance:
        status = att.Status
        if status.lower()  == 'week off':
            week_off_count += 1
        elif status.lower() == 'present':
            present_count += 1
        elif status.lower() == 'absent' or status is None:
            absent_count += 1
        elif status in leavelist:
            leave_counts[status] += 1

    total_working_days = week_off_count + present_count + sum(leave_counts.values())
    return week_off_count, present_count, absent_count, leave_counts, total_working_days

# ---- calculate_paid_days ----

def calculate_paid_days(total_working_days, total_days, total_no_days_in_month):
    if total_working_days>total_no_days_in_month:
        adjustment = len(total_days) - total_no_days_in_month
        return total_working_days - adjustment
    return total_working_days

# ---- days_in_selected_month

def days_in_selected_month(month_no, year):
    if month_no == 12:
        next_month = 1
        next_month_year = year + 1
    else:
        next_month = month_no + 1
        next_month_year = year
    start_date = datetime(year, month_no, 1)
    end_date = datetime(next_month_year, next_month, 1)
    return (end_date - start_date).days


# ------ Attendance_Lock_Post_API

def Attendance_Lock_Post_API(request):
    # if 'OrganizationID' not in request.session:
    #     return redirect(MasterAttribute.Host)
    
    # OrganizationID = request.session["OrganizationID"]

    try:
        data = json.loads(request.body)
    except Exception as e:
        return JsonResponse({"error": f"JSON decode failed: {str(e)}"}, status=400)
    
    OrganizationID = request.session["OrganizationID"]
    UserID = str(request.session["UserID"])


    EmployeeCode = data.get('emp')
    year = int(data.get('year'))
    month_no = int(data.get('month_no'))
    btn = data.get('btn')  # "Save" or "Lock"
    IsLock = btn == "Lock"
    lockid = data.get("lockid")
    # EmpName = data.get("empname")

    # print("Received Data:", data)
    # print("btn Type -----:", btn)
    # print("EmployeeCode:", EmployeeCode, "Month:", month_no, "Year:", year, "empname:",EmpName)

    month_name = calendar.month_name[month_no]

    # Correctly calculate the start and end dates
    start_date = get_previous_month_start_date(year, month_no)
    end_date = start_date.replace(day=28) + timedelta(days=4)
    end_date = end_date.replace(day=25)

    # # Generate total_days list
    total_days = [(start_date + timedelta(days=i)).strftime('%#d') for i in range((end_date - start_date).days + 1)]

    attendance = SalaryAttendance.objects.filter(EmployeeCode=EmployeeCode, OrganizationID=OrganizationID, IsDelete=False, Date__range=(start_date, end_date)).order_by('Date')
    leave_type = Leave_Type_Master.objects.filter(IsDelete=False, Is_Active=True)
    leavelist = [leave.Type for leave in leave_type]


    if lockid and lockid != '0':
        lock_obj = get_object_or_404(AttendanceLock, id=lockid, IsDelete=False)
        # print(len(total_days))
    if request.method == "POST":
        # btn = request.POST['btn']
        IsLock = btn == "Lock"

        week_off_count, present_count, absent_count, leave_counts, total_working_days = get_attendance_counts(attendance, leavelist)
        total_no_days_in_month = days_in_selected_month(month_no, year)
        pai_days = calculate_paid_days(total_working_days, total_days, total_no_days_in_month)
        #pai_days = len(total_days)
        # Adjust the paid days if there are 31 days from 26th to 25th, but the month has only 30 days
        # print(total_days)
        if pai_days>total_no_days_in_month:
            if len(total_days) == 30 and total_no_days_in_month == 31:
                pai_days += 1
            elif len(total_days) == 31 and total_no_days_in_month == 30:
                pai_days -= 1
            elif len(total_days) == 30 and total_no_days_in_month == 28:  # Adjust for non-leap year February
                pai_days -= 2
            elif len(total_days) == 31 and total_no_days_in_month == 29:  # Adjust for leap year February
                pai_days -= 1
        
        # print("the total_no_days_in_month is here ===========", total_no_days_in_month)

        if lockid and lockid != '0':
            # print("total_no_days_in_month  is here in update:", total_no_days_in_month)
            lock_obj.total_no_Days_in_month=total_no_days_in_month
            lock_obj.PaiDays = pai_days
            lock_obj.ModifyBy = UserID
            lock_obj.IsLock = IsLock
            lock_obj.save()
        else:
            # print("total_no_days_in_month  is here in create:", total_no_days_in_month)

            AttendanceLock.objects.create(
                OrganizationID=OrganizationID,
                CreatedBy=UserID,
                EmployeeCode=EmployeeCode,
                month=month_no,
                month_name=month_name,
                year=year,
                PaiDays=pai_days,
                total_no_Days_in_month=total_no_days_in_month,
                IsLock=IsLock
            )
            print("Object creation is trigered")

        # messages.success(request, "Saved Successfully" if btn == "Save" else "Locked Successfully")
        # return redirect(reverse('Employees_Payroll_List') + f'?year={year}&month_no={month_no}')
        return JsonResponse({"success": True, "message": f"{'Locked' if IsLock else 'Saved'} Successfully"})
    # return JsonResponse({"success": True, "message": f"{'Locked' if IsLock else 'Saved'} Successfully"})



#  ----- Employees_Payroll_List_API 

def Employees_Payroll_List_API(request):
    if 'OrganizationID' not in request.session:
        return JsonResponse({'error': 'Unauthorized'}, status=401)

    OrganizationID = request.session["OrganizationID"]
    UserID = str(request.session["UserID"])
    UserType = request.session["UserType"]
    current_date = datetime.now()

    year = int(request.GET.get('year', current_date.year))
    month_no = int(request.GET.get('month_no', current_date.month))

    # Step 1: Get all employees first
    emp_list = EmployeeDataSelectForSalary(OrganizationID, month_no, year)

    # Step 2: Get all EmployeeCodes from SalaryAttendance
    salary_attendance = SalaryAttendance.objects.filter(
        month=month_no,
        year=year,
        OrganizationID=OrganizationID,
        IsDelete=False,
        IsAttendanceMoved=True
    )

    # 3: Get only those EmployeeCodes that exist in SalaryAttendance
    salary_emp_codes = set(salary_attendance.values_list('EmployeeCode', flat=True))

    # 4: Filter emp_list to include only employees that exist in SalaryAttendance
    emp_list = [emp for emp in emp_list if emp['EmployeeCode'] in salary_emp_codes]

    # Step 5: Proceed with the rest of your existing logic
    previous_month = (datetime(year, month_no, 1) - timedelta(days=1)).strftime('%Y-%m-%d')
    next_month_date = min(
        datetime(year, month_no, 28) + timedelta(days=4),
        datetime.now().replace(day=1)
    )
    if next_month_date.month == month_no:
        next_month_date = next_month_date.replace(day=1)
    next_month = next_month_date.strftime('%Y-%m-%d')

    attendancelock = AttendanceLock.objects.filter(
        EmployeeCode__in=[emp['EmployeeCode'] for emp in emp_list],
        month=month_no,
        year=year,
        OrganizationID=OrganizationID,
        IsDelete=False,
    ).values('EmployeeCode', 'IsLock', 'id', 'IsFC_Verified', 'IsHR_Verified')

    generated_dict = defaultdict(lambda: {
        'generated': False,
        'id': 0,
    })

    lock_dict = defaultdict(lambda: {'IsLock': False, 'id': 0, 'IsFC_Verified': False, 'IsHR_Verified': False})

    for SA in salary_attendance:
        generated_dict[SA.EmployeeCode] = {
            'MovedToPayroll': SA.IsAttendanceMoved,
            'id': SA.id,
        }

    for lock in attendancelock:
        lock_dict[lock['EmployeeCode']] = {'IsLock': lock['IsLock'], 'id': lock['id'], 'IsFC_Verified': lock['IsFC_Verified'], 'IsHR_Verified': lock['IsHR_Verified']}

    for emp in emp_list:
        code = emp['EmployeeCode']
        slip_info = generated_dict[code]
        lock_info = lock_dict[code]

        emp['id'] = slip_info['id']
        emp['IsLock'] = lock_info['IsLock']
        emp['lockid'] = lock_info['id']
        emp['IsFC_Verified'] = lock_info['IsFC_Verified']
        emp['IsHR_Verified'] = lock_info['IsHR_Verified']

    return JsonResponse({
        'emps': emp_list,
        'current_month': datetime(year, month_no, 1).strftime('%Y-%m-%d'),
        'previous_month': previous_month,
        'next_month': next_month,
        'month_no': month_no,
        'year': year
    }, safe=False)


#  ---- Attendance_Lock_View_Get_Api
def Attendance_Lock_View_Get_Api(request):
    if 'OrganizationID' not in request.session:
        return JsonResponse({'error': 'Invalid session'}, status=401)

    OrganizationID = request.session["OrganizationID"]
    EmployeeCode = request.GET.get('emp')
    year = int(request.GET.get('year'))
    month_no = int(request.GET.get('month_no'))
    month_name = calendar.month_name[month_no]

    # emp_Details = EmployeeDataSelect(OrganizationID, EmployeeCode)

    # Calculate date range (26th to 25th logic)
    start_date = get_previous_month_start_date(year, month_no)
    end_date = start_date.replace(day=28) + timedelta(days=4)
    end_date = end_date.replace(day=26)
    
    # ------------(Before Organiztion Filter)-------------------
    # print("start_date::",  start_date)
    # print("end_date::",  end_date)
    # -------------------------------

    od = Organization_Details.objects.filter(OID=OrganizationID, IsDelete=False).first()
    if od and od.EndDate == 1:
        start_date = datetime(year, month_no, 1)
        last_day = calendar.monthrange(year, month_no)[1]
        end_date = datetime(year, month_no, last_day)

        # ------------(After Organiztion Filter)-------------------
        # print("start_date-2::",  start_date)
        # print("end_date-2::",  end_date)
        # # -------------------------------
        
    
    # Generate all dates in the range
    date_list = [start_date + timedelta(days=i) for i in range((end_date - start_date).days + 1)]

    # Create absent entries if missing
    # for date in date_list:
    #     if not SalaryAttendance.objects.filter(EmployeeCode=EmployeeCode, HotelID=OrganizationID, IsDelete=False, Date=date).exists():
    #         SalaryAttendance.objects.create(
    #             Date=date,
    #             Status="Absent",
    #             HotelID=OrganizationID,
    #             EmployeeCode=EmployeeCode
    #         )
    
    print("Employee Code::",  EmployeeCode)
    print("Organization ID::",  OrganizationID)
    print("start_date::",  start_date)
    print("end_date::",  end_date)
    
    # print("----------------------------------------------------------------")
    # EmployeeCode = "1000"
    # OrganizationID = "20180612060935"
    # Mystart_date = datetime(2025, 11, 1)
    # Myend_date = datetime(2025, 11, 30)
    # print("----------------------------------------------------------------")
    
    # print("Employee Code::",  EmployeeCode)
    # print("Organization ID::",  OrganizationID)
    # print("Mystart_date::",  Mystart_date)
    # print("Myend_date::",  Myend_date)

    # Fetch attendance records
    attendance = SalaryAttendance.objects.filter(
        EmployeeCode=EmployeeCode,
        HotelID=OrganizationID,
        IsDelete=False,
        Date__range=(start_date, end_date)
    ).order_by('Date')

    # Filter unique attendance per day, prioritize Present
    unique_attendance = []
    for date, records in groupby(attendance, key=attrgetter("Date")):
        records_list = list(records)
        present_record = next((r for r in records_list if r.Status == "Present"), None)
        selected_record = present_record if present_record else records_list[0]
        unique_attendance.append({
            'id': selected_record.id,  
            'date': selected_record.Date.strftime('%Y-%m-%d'),
            'status': selected_record.Status,
            'actual_status': selected_record.ActualStatus,
            'In_Time': selected_record.In_Time,
            'Out_Time': selected_record.Out_Time,
            'Duty_Hour': selected_record.Duty_Hour,
            'Remarks': selected_record.Remarks
        })

    # Leave types
    leave_type = Leave_Type_Master.objects.filter(IsDelete=False, Is_Active=True)
    leavelist = [leave.Type for leave in leave_type]

    return JsonResponse({
        'attendance_data': unique_attendance,
        'leave_types': leavelist,
        # 'employee_details': emp_Details,
        'month': month_name,
        'year': year
    })



#  ------------ Unlock Attendence

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.shortcuts import get_object_or_404
import calendar
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import get_object_or_404
import json
import calendar

@csrf_exempt
def Attendance_UnLock_Post_API(request):
    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "Only POST allowed"}, status=405)

    if 'OrganizationID' not in request.session:
        return JsonResponse({"status": "error", "message": "Session expired"}, status=403)

    try:
        data = json.loads(request.body)
        OrganizationID = request.session["OrganizationID"]
        UserID = str(request.session["UserID"])
        oid = data.get('oid')
        empid = data.get('empid')
        EmployeeCode = data.get('emp')
        year = int(data.get('year'))
        month_no = int(data.get('month_no'))
        lockid = data.get('lockid')
        btn_type = data.get('btn')

        # Optionally fetch employee details if needed
        # emp_Details = EmployeeDataSelect(OrganizationID, EmployeeCode)
        
        month_name = calendar.month_name[month_no]

        if lockid and lockid != '0':
            lock_obj = get_object_or_404(AttendanceLock, id=lockid, IsDelete=False)
            lock_obj.delete()

            Salary_Slip_V1.objects.filter(
                EmpID=empid,
                EmployeeCode=EmployeeCode,
                OrganizationID=oid,
                year=year,
                month=month_no,
                IsDelete=False
            ).delete()

            Delete_Per_Day_Salary(empid, oid, year, month_no)

        return JsonResponse({
            "status": "success",
            "message": f"Attendance unlocked for {EmployeeCode} ({month_name} {year})"
        })
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)}, status=500)





# ------------------ Bulk Lock Functionlity
import json
from django.http import JsonResponse


# ------------------ Bulk Lock Functionality
def Bulk_Lock_Attendence(request):
    if request.method == "POST":
        UserID = str(request.session.get("UserID", 0))
        OrganizationID = request.session.get("OrganizationID", 0)
        S_OID = OrganizationID

        # Read employees JSON array
        employees_json = request.POST.get("employees")
        if not employees_json:
            return JsonResponse({"error": "No employees provided"}, status=400)

        try:
            employees = json.loads(employees_json)   # list of dicts
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON format"}, status=400)

        # print("Employees received:", employees)

        results = []   # collect per-employee results

        for emp in employees:
            EmployeeCode = emp.get("emp_code")
            lockid = emp.get("lockid")
            year = emp.get("year")
            month_no = emp.get("month_no")
            IsLock = emp.get("islock")

            month_name = calendar.month_name[month_no]

            # print("Lock id is here::-->",lockid,"----------",type(lockid))

            # start_date = get_previous_month_start_date(year, month_no)
            # end_date = start_date.replace(day=28) + timedelta(days=4)
            # end_date = end_date.replace(day=25)

            # total_days = [(start_date + timedelta(days=i)).strftime('%#d')
            #               for i in range((end_date - start_date).days + 1)]

            # attendance = SalaryAttendance.objects.filter(
            #     EmployeeCode=EmployeeCode,
            #     OrganizationID=OrganizationID,
            #     IsDelete=False,
            #     Date__range=(start_date, end_date)
            # ).order_by('Date')

            # leave_type = Leave_Type_Master.objects.filter(IsDelete=False, Is_Active=True)
            # leavelist = [leave.Type for leave in leave_type]

            if lockid and lockid != 0:
                # print("Hello")
                lock_obj = get_object_or_404(AttendanceLock, id=lockid, IsDelete=False)
                lock_obj.ModifyBy = UserID
                lock_obj.IsLock = True
                lock_obj.save()
                results.append({"emp_code": EmployeeCode, "status": "updated"})
            else:
                AttendanceLock.objects.create(
                    OrganizationID=OrganizationID,
                    CreatedBy=UserID,
                    EmployeeCode=EmployeeCode,
                    month=month_no,
                    month_name=month_name,
                    year=year,
                    IsLock=True
                )
                results.append({"emp_code": EmployeeCode, "status": "created"})

        # Only return once, after processing all employees
        return JsonResponse({
            "month_no": month_no,
            "year": year,
            "OrganizationID": OrganizationID,
            "success": True,
            "message": f"Processed {len(employees)} employees",
        })









#  -------------------- Generate Salary Slip


def Generate_Salary_Employee_List_API(request):
    if 'OrganizationID' not in request.session:
        return JsonResponse({'error': 'Unauthorized'}, status=401)

    OrganizationID = request.session["OrganizationID"]
    UserID = str(request.session["UserID"])
    UserType = request.session["UserType"]
    current_date = datetime.now()

    year = int(request.GET.get('year', current_date.year))
    month_no = int(request.GET.get('month_no', current_date.month))

    # emp_list = EmployeeDataSelectForSalary(OrganizationID, month_no, year)

    previous_month = (datetime(year, month_no, 1) - timedelta(days=1)).strftime('%Y-%m-%d')
    next_month_date = min(
        datetime(year, month_no, 28) + timedelta(days=4),
        datetime.now().replace(day=1)
    )

    if next_month_date.month == month_no:
        next_month_date = next_month_date.replace(day=1)

    next_month = next_month_date.strftime('%Y-%m-%d')


    # Get full employee list
    emp_list = EmployeeDataSelectForSalary(OrganizationID, month_no, year)

    # Get all locked attendance records
    attendancelock = AttendanceLock.objects.filter(
        EmployeeCode__in=[emp['EmployeeCode'] for emp in emp_list],
        month=month_no,
        year=year,
        OrganizationID=OrganizationID,
        IsDelete=False,
        IsLock=True
    ).values('EmployeeCode', 'IsLock', 'id','IsGenerated')

    # Create a dictionary of locked employees
    lock_dict = {lock['EmployeeCode']: {'IsLock': lock['IsLock'], 'id': lock['id'], 'IsGenerated': lock['IsGenerated']} for lock in attendancelock}

    # Filter only employees whose attendance is locked
    locked_emps = []
    for emp in emp_list:
        code = emp['EmployeeCode']
        if code in lock_dict:
            emp['IsLock'] = True
            emp['lockid'] = lock_dict[code]['id']
            emp['get_approval_status_html'] = '<p style="color: #1ab394;">Attendance Locked</p>'
            emp['status'] = 'Locked'
            emp['IsGenerated'] = lock_dict[code]['IsGenerated']
            locked_emps.append(emp)


    return JsonResponse({
        'emps': locked_emps, 
        'current_month': datetime(year, month_no, 1).strftime('%Y-%m-%d'),
        'previous_month': previous_month,
        'next_month': next_month,
        'month_no': month_no,
        'year': year
    }, safe=False)



# Generate_Salary_Slip View
def Generate_Salary_Slip(request):
    if 'OrganizationID' not in request.session:
        return redirect(MasterAttribute.Host)
    else:
        print("Show Page Session")

    OrganizationID = request.session["OrganizationID"]
    # UserID = str(request.session["UserID"])
    # UserType = request.session["UserType"]
    current_date = datetime.now()


    year = request.GET.get('year')
    month_no = request.GET.get('month_no')

    if not year or not month_no:
        url = reverse("Generate_Salary_Slip")  
        return redirect(f"{url}?year={current_date.year}&month_no={current_date.month}")
    

    year = int(year)
    month_no = int(month_no)

    previous_month = datetime(year, month_no, 1) - timedelta(days=1)
    next_month = (datetime(year, month_no, 28) + timedelta(days=4)).replace(day=1)

    # Prevent future months
    if next_month > datetime.now().replace(day=1):
        next_month = datetime.now().replace(day=1)

    context = {
        'current_month': datetime(year, month_no, 1),
        'previous_month': previous_month,
        'next_month': next_month,
        'month_no': month_no,
        'year': year,
        'Session_OrganizationID ':OrganizationID
    }
    return render(request, "EMP_PAY/MoveToPayroll_Template/Generate_Salary_Slip.html", context)






# -- Per Day Salary Calculations

from datetime import date
import calendar
from decimal import Decimal
from HumanResources.models import Salary_Details_Effective, Salary_Detail_Master, SalaryTitle_Master, EmployeeBankInformationDetails, EmployeeIdentityInformationDetails, EmployeeWorkDetails, EmployeeEmergencyInformationDetails, EmployeePersonalDetails
import math
import calendar
from decimal import Decimal
from datetime import date, datetime, timedelta
from django.http import JsonResponse
from decimal import Decimal, InvalidOperation
import traceback
from django.db import transaction


# Bulk salary generation process
def Bulk_Generate_Salary_Process(request):
    if request.method == "POST":
        UserID = str(request.session.get("UserID", 0))
        OrganizationID = request.session.get("OrganizationID", 0)
        S_OID = OrganizationID

        # Read employees JSON array
        employees_json = request.POST.get("employees")
        if not employees_json:
            return JsonResponse({"error": "No employees provided"}, status=400)

        try:
            employees = json.loads(employees_json)  # list of dicts
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON format"}, status=400)

        results = []
        failed = []
        success = []

        for emp in employees:
            EmpCode = emp.get("emp_code")
            EmpID = emp.get("empid")
            EmpName = emp.get("empname")
            designation = emp.get("designation")
            department = emp.get("department")
            Doj = emp.get("doj")
            OID = emp.get("oid")
            year = emp.get("year")
            month_no = emp.get("month_no")

            month_no_int = int(month_no)
            Year_int = int(year)
            month_name = calendar.month_name[month_no_int]

            try:
                # Run helper (may raise exceptions)
                HelperData = Helper_process_daily_earning(
                    OID, EmpID, EmpCode, EmpName, Year_int, month_no_int,
                    department, designation, Doj, UserID
                )

                ResponseStatus = "Success"
                ResponseMessage = "Processed successfully"

                results.append({
                    "EmpCode": EmpCode,
                    "EmpName": EmpName,
                    "status": "success",
                    "message": ResponseMessage
                })
                success.append(EmpCode)

            except Exception as e:
                # Record failure but continue
                ResponseStatus = "Error"
                ResponseMessage = str(e)

                results.append({
                    "EmpCode": EmpCode,
                    "EmpName": EmpName,
                    "status": "error",
                    "message": ResponseMessage
                })
                failed.append({
                    "EmpCode": EmpCode,
                    "EmpName": EmpName,
                    "Error": ResponseMessage
                })

            # Delete old record if it exists
            Response_Data_Generate_Slip.objects.filter(
                Month_No=month_no,
                Year=year,
                Emp_Code=EmpCode,
                Emp_ID=EmpID,
                Hotel_Id=OID
            ).delete()

            # Insert new record
            Response_Data_Generate_Slip.objects.create(
                Month_No=month_no,
                Year=year,
                Emp_Name=EmpName,
                Emp_Code=EmpCode,
                Emp_ID=EmpID,
                Response_Status=ResponseStatus,
                Response_Message=ResponseMessage,
                Hotel_Id=OID,
                OrganizationID=OrganizationID,
                CreatedBy=UserID,
                CreatedDateTime=timezone.now(),
            )

        print("I am at bulk salary generation")
        return JsonResponse({
            "month_no": month_no,
            "year": year,
            "OID": OID,
            "success": True,
            "total_employees": len(employees),
            "processed_successfully": len(success),
            "failed_count": len(failed),
            "failed_employees": failed,
            "results": results,
            "message": f"Processed {len(success)} successfully, {len(failed)} failed."
        }, status=200)




def process_daily_earning(request):
    try:
        if request.method != "POST":
            return JsonResponse({"status": "error", "message": "Only POST allowed"}, status=405)

        data = json.loads(request.body)

        OID = data.get('oid')
        EmpID = data.get('EmpID')
        EmpCode = data.get('empCode') or 0
        EmpName = data.get('empname') or ''
        year = int(data.get('year'))
        month = int(data.get('month'))

        department = data.get('department') or ''
        designation = data.get('designation') or ''
        Doj = data.get('Doj') or ''
        UserID = request.session.get("UserID", 0)

        # print("the Empcode is here*------------*", EmpCode)
        print("------- Step-1 ------")

        Helper_process_daily_earning(OID, EmpID, EmpCode, EmpName, year, month, department, designation, Doj, UserID)
        return JsonResponse({"status": "success"})
    
        # return JsonResponse({
        #     "status": "success",
        #     "message": "Salary computed with split periods"
        # }, status=200)
    
    except ValueError as e:
        return JsonResponse({"status": "error", "message": str(e)}, status=404)
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)}, status=500)


@transaction.atomic
def Helper_process_daily_earning(OID, EmpID, EmpCode, EmpName, year, month, department='', designation='', Doj='', UserID=0):
    print("Helper_process_daily_earning hit")

    od = Organization_Details.objects.filter(OID=OID, IsDelete=False).first()

    if od:
        if od.EndDate == 1:  
            # Cycle = 1st → Last Day of Month
            month_start = datetime(year, month, 1).date()
            last_day = calendar.monthrange(year, month)[1]
            month_end = datetime(year, month, last_day).date()

        elif od.EndDate == 31:
            # Cycle = 1st → 31st (normal calendar month)
            month_start = datetime(year, month, 1).date()
            last_day = calendar.monthrange(year, month)[1]  # handles Feb (28/29), April (30), etc.
            month_end = datetime(year, month, last_day).date()

        else:
            # Default cycle: 26th prev month → EndDate current month
            month_start = get_previous_month_start_date(year, month).date()
            month_end = datetime(year, month, od.EndDate).date()
    else:
        # Fallback to default cycle if org details not found
        month_start = get_previous_month_start_date(year, month)
        month_end = datetime(year, month, 25)


    # total_days_in_month = (month_end - month_start).days + 1  # +1 to include both start and end
    # total_days_in_month = (month_end - month_start).days   # Not Include include both start and end
    total_days_in_month = calendar.monthrange(year, month)[1]   
    # print("Total Days in Month:", total_days_in_month)

    # print("Data before Effective records")
    effective_records = Salary_Details_Effective.objects.filter(
        EmpID=EmpID,
        OrganizationID=OID,
        IsDelete=False,
        EffectiveFrom__lte=month_end
    ).order_by('EffectiveFrom')
    
    print("Helper_process_daily_earning hit--- 1.0")
    

    
    # print("I just want to exeecute after Effective records", effective_records)


    if not effective_records.exists():
        print("No effective salary found for this month")
        raise ValueError("No effective salary found for this month")

    print("Helper_process_daily_earning hit--- 2.0")

    # Loop through each effective salary
    for idx, effective_obj in enumerate(effective_records):
        effective_from = effective_obj.EffectiveFrom
        next_effective_from = effective_records[idx + 1].EffectiveFrom if idx + 1 < len(effective_records) else None
        
        print("Helper_process_daily_earning hit--- 2.2")

        # Determine date range
        start_date = effective_from
        # print("Start Date:", start_date)
        end_date_str = next_effective_from - timedelta(days=1) if next_effective_from else month_end
        end_date = end_date_str.date() if isinstance(end_date_str, datetime) else end_date_str
        print("Helper_process_daily_earning hit--- 2.4")
        

        # Skip if outside current month
        if start_date < month_start:
            start_date = month_start
        if end_date > month_end:
            end_date = month_end
            
        print("Helper_process_daily_earning hit--- 2.5")

        # days_in_period = (end_date - start_date).days + 1
        # days_in_period = (end_date - start_date).days   # Not Include include both start and end
        # days_in_period = (end_date - start_date).days + 1

        # print(f"Processing from {start_date} to {end_date} ({days_in_period} days)")

        # Fetch salary details
        SalaryTitles = SalaryTitle_Master.objects.filter(IsDelete=False, HotelID=OID, Type="A").order_by('TypeOrder', 'TitleOrder')
        Salary_Detail_Data = Salary_Detail_Master.objects.filter(
            Effective=effective_obj, IsDelete=False, OrganizationID=OID
        ).values('Salary_title_id', 'Permonth', 'Perannum')

        salary_map = {s['Salary_title_id']: s for s in Salary_Detail_Data}

        salary_values = {}
        for salary in SalaryTitles:
            detail = salary_map.get(salary.id)
            permonth = Decimal(detail['Permonth']) if detail else Decimal('0')

            salary_values[salary.Title] = {
                "id": salary.id,
                "permonth": permonth,
                "Type": salary.Type,
                "TypeOrder": salary.TypeOrder,
                "TitleOrder": salary.TitleOrder,
            }

        total_days = total_days = Decimal(calendar.monthrange(year, month)[1])
        print("Helper_process_daily_earning hit--- 3.0")
        
        
        
        # Another Trial (2nd) -------------------------------------------
        full_absent = SalaryAttendance.objects.filter(
            EmpID=EmpID,
            Date__range=(start_date, end_date),
            OrganizationID=OID,
            IsDelete=False,
            Status='Absent'
        ).count()
        
        half_day = SalaryAttendance.objects.filter(
            EmpID=EmpID,
            Date__range=(start_date, end_date),
            OrganizationID=OID,
            IsDelete=False,
            Status='Half Day Present'
        ).count()
        
        absent_days = Decimal(full_absent) + (Decimal(half_day) * Decimal("0.5"))

        payable_days = Decimal(total_days) - absent_days
        print("Helper_process_daily_earning hit--- 4.0")

        if payable_days < 0:
            payable_days = Decimal("0")

        print(f"Start Date: {start_date}, End Date:{end_date}")
        print("Total Days is here:", total_days)
        print("half_day is here:", half_day)
        print("absent_days is here:", absent_days)
        print("payable_days is here:", payable_days)
        
        # Trial End -------------------------------------------
        

        # Fetch Salary Titles (Type = A)
        SalaryTitles = SalaryTitle_Master.objects.filter(
            IsDelete=False,
            HotelID=OID,
            Type="A"
        ).order_by('TypeOrder', 'TitleOrder')

        Salary_Detail_Data = Salary_Detail_Master.objects.filter(
            Effective=effective_obj,
            IsDelete=False,
            OrganizationID=OID
        ).values('Salary_title_id', 'Permonth')

        salary_map = {s['Salary_title_id']: Decimal(s['Permonth']) for s in Salary_Detail_Data}

        # FINAL MONTHLY CALCULATION (ONLY ONCE)
        print("------- FINAL MONTHLY CALCULATION (ONLY ONCE) (Entring In Salary Earning) ------")
        
        for title in SalaryTitles:

            permonth = salary_map.get(title.id, Decimal("0"))

            final_amount = (
                permonth / total_days * payable_days
                if payable_days > 0 else Decimal("0")
            )

            final_amount = final_amount.quantize(Decimal("0.01"))

            Salary_Earning_Details.objects.update_or_create(
                # SalaryAttendance=0,
                EmpID=EmpID,
                EmpCode=EmpCode,
                SalaryTitleID=title.id,
                OrganizationID=OID,
                year=year,
                month=month,
                defaults={
                    "SalaryTitle": title.Title,
                    "Amount": final_amount,
                    "Type": title.Type,
                    "TypeOrder": title.TypeOrder,
                    "TitleOrder": title.TitleOrder,
                    "CreatedBy": UserID,
                }
            )
            
        print("entring in Update_Earning_Bonus")
        Update_Earning_Bonus(month, year, OID, EmpID)


        print("entring in get_Deduction_salary_components")
        get_Deduction_salary_components(EmpID, OID, year, month, UserID)

    # print("------- Step-2 (Done) - and - 3 (Started) ------")
    generate_salary_slip_v1(EmpID,EmpCode,EmpName,designation,department, OID, year, month,total_days,absent_days,payable_days,UserID, Doj)
    # print("------- Step-9 ------ generate_salary_slip_v1")
    # print("------------- Successfully Saved --------------")

    # return JsonResponse({
    #     "status": "success",
    #     "message": "Salary computed with split periods"
    # }, status=200)
    return True


# Get Deduction salary components
def get_Deduction_salary_components(EmpID, OID, year, month, UserID):
    # print("----- Step-5 ----- get_Deduction_salary_components ---")

    # Get the latest effective salary record
    latest_effective = (
        Salary_Details_Effective.objects
        .filter(EmpID=EmpID, OrganizationID=OID, IsDelete=False)
        .order_by('-EffectiveFrom')
        .first()
    )

    if not latest_effective:
        print("No effective salary found for this employee")
        raise ValueError("No effective salary found for this employee")
        # return JsonResponse({
        #     "status": "error",
        #     "message": "No effective salary found for this employee"
        # }, status=404)
        # return None

    # Get any attendance object (for relational link)
    attendance = (
        SalaryAttendance.objects
        .filter(EmpID=EmpID, month=month, year=year, OrganizationID=OID, IsDelete=False)
        .first()
    )

    # print(f"attendance obj is params data here: Empid: {EmpID}, month: {month}, year: {year}, OID: {OID}")
    # print("attendance obj is here:", attendance)
    if attendance is None:
        print("Attendance record not found for this month")
        raise ValueError("Attendance record not found for this month")
        # return JsonResponse({
        #     "status": "error",
        #     "message": "Attendance record not found for this month"
        # }, status=404)

    # Get Fixed Salary titles (Type = 'B')
    SalaryTitles = (
        SalaryTitle_Master.objects
        .filter(IsDelete=False, HotelID=OID, Type='B')
        .order_by('TypeOrder', 'TitleOrder')
    )

    # Get Fixed Salary titles (Type = 'A')
    SalaryTitles_A = (
        SalaryTitle_Master.objects
        .filter(IsDelete=False, HotelID=OID, Type='A')
        .order_by('TypeOrder', 'TitleOrder')
    )

    # Get Fixed Salary titles (Type = 'C')
    SalaryTitles_C = (
        SalaryTitle_Master.objects
        .filter(IsDelete=False, HotelID=OID, Type='C')
        .order_by('TypeOrder', 'TitleOrder')
    )

    # Get Salary Details under this effective record
    Salary_Detail_Data = (
        Salary_Detail_Master.objects
        .filter(Effective=latest_effective, IsDelete=False)
        .values('Salary_title_id', 'Permonth', 'Perannum')
    )

    salary_map = {s['Salary_title_id']: s for s in Salary_Detail_Data}

    # Loop over each fixed salary title and store
    for salary in SalaryTitles:
        detail = salary_map.get(salary.id)
        permonth = Decimal(detail['Permonth']) if detail else Decimal('0')

        # Create or update entry in Salary_Fixed_Details
        Salary_Deduction_Details.objects.update_or_create(
            SalaryAttendance=attendance,
            SalaryTitle=salary.Title,
            SalaryTitleID=salary.id,
            OrganizationID=OID,
            defaults={
                "Amount": permonth,
                "Type": salary.Type,   
                "TypeOrder": salary.TypeOrder,  
                "TitleOrder": salary.TitleOrder,   
                "CreatedBy": UserID,
                "Date": attendance.Date if attendance else None
            }
        )

    # Loop over each fixed salary title and store
    for salary in SalaryTitles_A:
        detail = salary_map.get(salary.id)
        permonth = Decimal(detail['Permonth']) if detail else Decimal('0')

        # Create or update entry in Salary_Fixed_Details
        Salary_Fixed_Details.objects.update_or_create(
            SalaryAttendance=attendance,
            SalaryTitle=salary.Title,
            SalaryTitleID=salary.id,
            OrganizationID=OID,
            defaults={
                "Amount": permonth,
                "Type": salary.Type, 
                "TypeOrder": salary.TypeOrder,
                "TitleOrder": salary.TitleOrder,   
                "CreatedBy": UserID,
                "Date": attendance.Date if attendance else None
            }
        )

    # Loop over each fixed salary title and store
    for salary in SalaryTitles_C:
        detail = salary_map.get(salary.id)
        permonth = Decimal(detail['Permonth']) if detail else Decimal('0')

        # Create or update entry in Salary_Fixed_Details
        Salary_Company_Contribution_Details.objects.update_or_create(
            SalaryAttendance=attendance,
            SalaryTitle=salary.Title,
            SalaryTitleID=salary.id,
            OrganizationID=OID,
            defaults={
                "Amount": permonth,
                "Type": salary.Type, 
                "TypeOrder": salary.TypeOrder,
                "TitleOrder": salary.TitleOrder,   
                "CreatedBy": UserID,
                "Date": attendance.Date if attendance else None
            }
        )

    # print("Fixed salary details saved successfully for employee:", EmpID)
    return True


# Calculate Bonus
# def Calculate_Bonus(fixed_basic, Basic_Earning):
#     """
#     Formula Used:
#         If Basic_Earning < 21000 → Bonus = Basic_Earning × 8.33%
#         Else → Bonus = 0.00
#     """

#     if fixed_basic < 21000:
#         bonus = (Basic_Earning * 8.33) / 100
#     else:
#         bonus = 0.00

#     return round__Value_decimal(bonus)


from decimal import Decimal

def Calculate_Bonus(fixed_basic, Basic_Earning, OID):
    """
    Formula Used:
        If fixed_basic < 21000 → Bonus = Basic_Earning × 8.33%
        Else → Bonus = 0.00
    """
    if OID == '501':
        if fixed_basic < Decimal("21000"):
            bonus = (Basic_Earning * Decimal("8.33")) / Decimal("100")
        else:
            bonus = Decimal("0.00")
    else:
        bonus = Decimal("0.00")

    return round__Value_decimal(bonus)



# Calculate_Employe_r_PF
def Calculate_Employe_r_PF(Data):
    """
    Formula Used 
        =IF((Basic+Conveyance Allowance+Special Allowance)>=15000, 1950, (BASIC+Conveyance Allowance+Special Allowance)*13%) - employer contribution on earned salary -- (Exclude HRA)
    """
    
    if Data >= 15000:
        EmployeerPF = 1950
    else:
        EmployeerPF = (Data * 13) / 100

    # return EmployeerPF
    return round__Value_decimal(EmployeerPF)


# Calculate Employee PF
def Calculate_EmployeePF(Data):
    """
    Formula Used 
        =IF((Basic+Conveyance Allowance+Special Allowance)>=15000, 1800, (Basic+Conveyance Allowance+Special Allowance)*12%) - Employee PF contribution on Earned Salary -- (Exclude HRA)
    """

    if Data >= 15000:
        Employee_PF = 1800
    else:
        Employee_PF = (Data * 12) / 100 
        if Employee_PF > 1800:
            Employee_PF = 1800

    return round__Value_decimal(Employee_PF)


from decimal import Decimal, ROUND_HALF_UP, ROUND_UP

# Calculate ESIC
# def Calculate_ESIC(OID, gross_salary, total_earning):
# def Calculate_ESIC(OID, gross_salary, total_earning):
def Calculate_ESIC(OID, fixed_basic):
    """
        Formula Used 
            ESIC (Employee Contribution) =IF(Actual Gross<=21000,ROUNDUP(Earned Gross*0.75%,0),0) - Employee Contribution - Employee Contribution - ESIC
            CompanyContributionToESIC (Employee'r Contribution) =IF(Actual Gross<=21000,ROUNDUP(Earned Gross*3.25%,0),0) - Employee'r Contribution - ESIC
    """

    OrgConfig = Organization_Details.objects.filter(OID=OID, IsDelete=False).first()

    if not OrgConfig or not OrgConfig.IsESICCalculate:
        return {'ESIC': Decimal('0'), 'CompanyContributionToESIC': Decimal('0.00')}

    # Excel logic: IF(Actual Gross <= 21000)
    if fixed_basic > 21000:
        return {'ESIC': Decimal('0'), 'CompanyContributionToESIC': Decimal('0.00')}

    # Calculate raw values
    ESIC_raw = (fixed_basic * Decimal('0.75') / Decimal('100'))
    Company_raw = (fixed_basic * Decimal('3.25') / Decimal('100'))

    # ROUNDUP to next whole rupee
    ESIC = round_decimal(ESIC_raw)
    CompanyContributionToESIC = round_decimal(Company_raw)

    return {
        'ESIC': ESIC,
        'CompanyContributionToESIC': CompanyContributionToESIC
    }

# Calculate PT
def Calculate_PT(OrganizationID, Total_Earning, Gender, month_no):
    
    print("------------------- (PT Calcualtions) -------------------------------")
    print(f"OrganizationID:{OrganizationID}, Total_Earning:{Total_Earning}, Gender:{Gender},month_no:{month_no}")
    print("------------------- (PT Ends) -------------------------------")

    PT = 0
    OID = str(OrganizationID)
    Total_Earning_Year = Total_Earning * 12 or 0 

    if OID in ['401','501','601','1501','20180612060935','2130']:
        PT = 200 if Total_Earning >= 12001 else 0
        print("We are in PT Calculations: and the pt is here", PT)

    elif OID == '1101':
        if Total_Earning <= 5999:
            PT = 0
        elif Total_Earning > 6000 and Total_Earning <= 8999:
            PT = 80
        elif Total_Earning > 9000 and Total_Earning <= 11999:
            PT = 150
        elif Total_Earning > 12000:
            PT = 200
        else:
            PT = 0

    elif OID == '1901':
        if Total_Earning <= 15000:
            PT = 0
        elif Total_Earning > 15000 and Total_Earning <= 20000:
            PT = 150
        elif Total_Earning > 20000:
            PT = 200
        else:
            PT = 0

    elif OID == '2010':
        if Total_Earning_Year <= 250000:
            PT = 0
        elif Total_Earning > 250000:
            PT = 200
        else:
            PT = 0
            
    elif OID == '1401':
        if Total_Earning_Year <= 225000:
            PT = 0
        elif Total_Earning_Year > 225000 and Total_Earning_Year <= 300000:
            PT = 125
        elif Total_Earning_Year > 300000 and Total_Earning_Year <= 400000:
            if month_no==3:
                PT = 174
            else:
                PT = 166
        elif Total_Earning_Year > 400000:
            if month_no==3:
                PT = 212
            else:
                PT = 208
        else:
            PT = 0
        
    elif OID in ['2140','2020']:
        if Gender=="Female":
            # print("the Earned_Basic is already", Earned_Basic)
            if Total_Earning <25000:
                PT=0
            elif Total_Earning>=25000:
                if month_no==2:
                    PT=300
                else:
                    PT=200
            # print("the PT is already", PT)

        else: 
            if Total_Earning<=7500:
                PT=0
            elif Total_Earning>7500 and Total_Earning <= 10000:
                PT=175
            elif Total_Earning>10000:
                if month_no==2:
                    PT=300
                else:
                    PT=200
                    
    elif OID == '2100':  # Jharkhand (Ranchi)
        
        if Total_Earning <= 25000:
            PT = 0
        elif Total_Earning <= 41666:
            PT = 100
        elif Total_Earning <= 66666:
            PT = 150
        elif Total_Earning <= 83333:
            PT = 175
        else:
            PT = 208


    return PT


# Update_Earning_Bonus
def Update_Earning_Bonus(Month, Year, OID, EmpID):

    # Get earned basic
    Basic_Salary_Earned = (
        Salary_Earning_Details.objects
        .filter(
            month=Month,
            year=Year,
            OrganizationID=OID,
            EmpID=EmpID,
            SalaryTitle="Basic",
            IsDelete=False
        )
        .aggregate(total=Sum('Amount'))['total'] or Decimal("0.00")
    )

    # Get fixed basic
    fixed_basic = (
        Salary_Fixed_Details.objects
        .filter(
            SalaryAttendance__EmpID=EmpID,
            SalaryAttendance__OrganizationID=OID,
            SalaryAttendance__year=Year,
            SalaryAttendance__month=Month,
            SalaryTitle="Basic",
            IsDelete=False
        )
        .aggregate(total=Sum('Amount'))['total'] or Decimal("0.00")
    )

    # Calculate bonus
    Calculate_Bonus_Value = Calculate_Bonus(
        Decimal(fixed_basic),
        Decimal(Basic_Salary_Earned),
        OID
    )

    # ONLY UPDATE (no create)
    Salary_Earning_Details.objects.filter(
        month=Month,
        year=Year,
        OrganizationID=OID,
        EmpID=EmpID,
        SalaryTitle="Bonus",
        IsDelete=False
    ).update(Amount=Calculate_Bonus_Value)

    return True


#  Genearte Salary Slip
from django.db.models import Sum
def generate_salary_slip_v1(emp_id,EmpCode,EmpName,department,designation, oid, year, month,total_days,absent_days, payable_days, user_id, Doj):

    BankInfo = EmployeeBankInformationDetails.objects.filter(EmpID=emp_id, OrganizationID=oid, IsDelete=False).first()
    EmpIdentity_Info = EmployeeIdentityInformationDetails.objects.filter(EmpID=emp_id, OrganizationID=oid, IsDelete=False).first()
    EmpWork_Info = EmployeeWorkDetails.objects.filter(EmpID=emp_id, OrganizationID=oid, IsDelete=False, IsSecondary=False).first()
    Personal_Info = EmployeePersonalDetails.objects.filter(EmpID=emp_id, OrganizationID=oid, IsDelete=False, IsEmployeeCreated=True).first()
    Emergency_Info = EmployeeEmergencyInformationDetails.objects.filter(EmpID=emp_id, OrganizationID=oid, IsDelete=False).first()

    # Check if slip already exists
    existing_slip = Salary_Slip_V1.objects.filter(
        EmpID=emp_id,
        EmployeeCode=EmpCode,
        OrganizationID=oid,
        year=year,
        month=month,
        IsDelete=False
    ).first()

    # Only create if not exists
    if not existing_slip:
        slip = Salary_Slip_V1.objects.create(
            EmpID=emp_id,
            EmployeeCode=EmpCode,
            # DOJ=EmpWork_Info.DateofJoining if EmpWork_Info else Doj,
            DOJ = EmpWork_Info.DateofJoining.strftime("%d-%m-%Y") if EmpWork_Info else Doj.strftime("%d-%m-%Y"),
            Emp_Name=EmpName,
            Designation=EmpWork_Info.Designation if EmpWork_Info else designation,
            Department=EmpWork_Info.Department if EmpWork_Info else department,
            EmpStatus=EmpWork_Info.EmpStatus if EmpWork_Info else department,
            OrganizationID=oid,
            year=year,
            month=month,
            month_name=calendar.month_name[month],
            Total_No_Days_In_Month=total_days,
            Present_Days=payable_days,
            Absent_Days=absent_days,

            Pan_No=EmpIdentity_Info.PANNo if EmpIdentity_Info else '',
            Bank_Account_Number=BankInfo.BankAccountNumber if BankInfo else '',
            Bank_Name=BankInfo.NameofBank if BankInfo else '',
            Bank_IFSC_Code=BankInfo.IFSCCode if BankInfo else '',
            Bank_Branch=BankInfo.BankBranch if BankInfo else '',
            Provident_Fund_Number=Emergency_Info.ProvidentFundNumber if Emergency_Info else '',
            ESIC_Number=Emergency_Info.ESINumber if Emergency_Info else '',
            IsGenerated=True,
            CreatedBy=user_id,
        )
        Gender = Personal_Info.Gender if Personal_Info else ''

        Calculate_Salary_Slip_Data(emp_id, oid, year, month, Gender)

        existing_Lock = AttendanceLock.objects.filter(
            # EmpID=emp_id,
            EmployeeCode=EmpCode,
            OrganizationID=oid,
            year=year,
            month=month
        ).first()
        
        if existing_Lock:
            existing_Lock.IsGenerated = True   
            existing_Lock.save(update_fields=["IsGenerated"])
    else:
        slip = existing_slip
        # print("Salary slip already exists, not creating a new one.")

    return slip



from decimal import Decimal, ROUND_DOWN
import math
from num2words import num2words
from django.db.models import Sum


def Show_Salary_Slip_PDF(request, EmpID, OID, Year, Month, Action=False):

    # print(f"action:{Action}")

    Action if Action else 'View' 
    
    # === Fetch Basic Slip Information ===
    slip = Salary_Slip_V1.objects.filter(
        EmpID=EmpID,
        OrganizationID=OID,
        year=Year,
        month=Month,
        IsDelete=False
    ).first()

    if not slip:
        return render(request, 'payslip_not_found.html')

    # === Fixed Salary ===
    fixed_details = list(
        Salary_Fixed_Details.objects
        .filter(
            SalaryAttendance__EmpID=EmpID,
            SalaryAttendance__OrganizationID=OID,
            SalaryAttendance__year=Year,
            SalaryAttendance__month=Month,
            IsDelete=False
        )
        .values('SalaryTitle')
        .annotate(total=Sum('Amount'))
        .order_by('TitleOrder')
    )

    fixed_details = fixed_details[:-1]
    Fixed = {item['SalaryTitle']: math.ceil(float(item['total'])) for item in fixed_details}
    # print("the fixed details are here::--->",Fixed)
    # Fixed = {item['SalaryTitle']: float(item['total']) for item in fixed_details}

    # === Earnings ===
    earning_details = list(
        Salary_Earning_Details.objects
        .filter(
            month=Month,
            year=Year,
            OrganizationID=OID,
            EmpID=EmpID,
            IsDelete=False
        )
        .values('SalaryTitle')
        .annotate(total=Sum('Amount'))
        .order_by('TitleOrder')
    )
    # earning_details = list(earning_details[:-1])
    earning_details = earning_details[:-1]
    # Earning = {item['SalaryTitle']: math.ceil(float(item['total'])) for item in earning_details}
    # Earning = {item['SalaryTitle']: float(item['total']) for item in earning_details}
    # Earning = {item['SalaryTitle']: Decimal(item['total']) for item in earning_details}

    Earning = {
        item['SalaryTitle']: Decimal(item['total']).quantize(Decimal('0.00'), rounding=ROUND_DOWN)
        for item in earning_details
    }


    
    # Earning_Show = {
    #     item['SalaryTitle']: f"{item['total']:.2f}"
    #     for item in earning_details
    # }
    Earning_Show = {
        item['SalaryTitle']: float(f"{item['total']:.2f}")
        for item in earning_details
    }

    # === Deductions ===
    deduction_details = list(
        Salary_Deduction_Details.objects
        .filter(
        SalaryAttendance__EmpID=EmpID,
        SalaryAttendance__OrganizationID=OID,
        SalaryAttendance__year=Year,
        SalaryAttendance__month=Month,
        IsDelete=False
        )
        .values('SalaryTitle')
        .annotate(total=Sum('Amount'))
        .order_by('TitleOrder')
    )
    deduction_details = deduction_details[:-1]
    Deductions = {item['SalaryTitle']: math.ceil(float(item['total'])) for item in deduction_details}

    # print("=== ----------- ===================== -------------------")
    # print("=== ----------- ===================== -------------------")
    # for item in deduction_details:
    #     title = item['SalaryTitle']
    #     total = math.ceil(float(item['total']))

    #     print(title, total)   # or use it anywhere

    # print("=== ----------- ===================== -------------------")
    # print("=== ----------- ===================== -------------------")


    # === Totals ===
    total_fixed = sum(Fixed.values())
    # print("the total fixed is here::--->", total_fixed)
    # total_earning = sum(Earning.values())
    total_earning = sum(Earning.values(), Decimal("0.00"))
    total_deduction = sum(Deductions.values())
    net_pay = Decimal(total_earning - total_deduction).quantize(Decimal('0.01'))
    net_pay_int = int(net_pay)
    net_pay_in_words = num2words(net_pay_int, lang='en_IN').replace(',', '').title()
    net_pay_in_words = f"{net_pay_in_words} Rupees Only"


    # === Organization Details ===
    org_Details  =  OrganizationMaster.objects.get(OrganizationID=OID)
    
    month_name = calendar.month_name[Month]

    # === Send to Template ===
    context = {
        'slip': slip,
        'Fixed': Fixed,
        'Earning': Earning_Show,
        # 'Earning': Earning,
        'Deductions': Deductions,
        'total_fixed': total_fixed,
        'total_earning': total_earning,
        'total_deduction': total_deduction,
        'month_name': calendar.month_name[Month],
        'net_pay_in_words': net_pay_in_words,
        'org_Details': org_Details,
        'net_pay': net_pay,
        'Action': Action,
        'Month': Month,
        'month_name': month_name,
        'year': Year,
    }

    return render(request, "EMP_PAY/MoveToPayroll_Template/Generate_Salary_Slip.html", context)

# Previous One --------------->
def Calculate_Salary_Slip_Data(EmpID, OID, Year, Month, Gender):
    print("Entring in Calculate_Salary_Slip_Data(CSSD)")
    print("(CSSD Step 1")
    # === Fixed Salary ===
    fixed_details = list(
        Salary_Fixed_Details.objects
        .filter(
            SalaryAttendance__EmpID=EmpID,
            SalaryAttendance__OrganizationID=OID,
            SalaryAttendance__year=Year,
            SalaryAttendance__month=Month,
            IsDelete=False
        )
        .values('SalaryTitle', 'TitleOrder')
        .annotate(total=Sum('Amount'))
        .order_by('TitleOrder')
    )
    fixed_details = fixed_details[:-1]
    Fixed = {item['SalaryTitle']: float(item['total']) for item in fixed_details}

    print("(CSSD Step 2")

    # === Earnings ===
    earning_details = list(
        Salary_Earning_Details.objects
        .filter(
            month=Month,
            year=Year,
            OrganizationID=OID,
            EmpID=EmpID,
            IsDelete=False
        )
        .values('SalaryTitle', 'TitleOrder')
        .annotate(total=Sum('Amount'))
        .order_by('TitleOrder')
    )
    earning_details = earning_details[:-1]
    Earning = {item['SalaryTitle']: float(item['total']) for item in earning_details}
    print("(CSSD Step 3")

    # === Totals ===
    # total_earning = Decimal(math.ceil(sum(Earning.values())))
    # total_earning = Decimal(math.ceil(sum(Earning.values())))
    # total_earning = Decimal(sum(Earning.values()))
    # total_Fixed = Decimal(sum(Fixed.values()))
    total_earning = Decimal(str(sum(Earning.values()))).quantize(Decimal("0.00"))
    total_Fixed = Decimal(str(sum(Fixed.values()))).quantize(Decimal("0.00"))

    # total_deduction = Decimal(sum(Deductions.values()))
    print("(CSSD Step 4")

    # --- Fetch required values for recalculation ---
    Basic_Salary_Earned = Decimal(Earning.get("Basic", 0))
    HRA_Earned = Decimal(Earning.get("HRA", 0))
    fixed_basic = Decimal(Fixed.get("Basic", 0))
    fixed_gross = Decimal(sum(Fixed.values()))
    
    Data_PF_Calculation = total_earning - HRA_Earned
    print("(CSSD Step 4.4")

    # Gender = 'Female' 
    month_no = int(Month)

    # === Call your calculation functions ===
    Calculated_PT = Calculate_PT(OID, total_earning, Gender, month_no)
    print("(CSSD Step 4.5")
    
    Calculated_Employee_PF = Calculate_EmployeePF(Data_PF_Calculation)
    print("(CSSD Step 4.6")
    
    Calculated_Employer_PF = Calculate_Employe_r_PF(Data_PF_Calculation)
    print("(CSSD Step 4.7")
    
    # Calculate_Bonus_Value = Calculate_Bonus(fixed_basic, Basic_Salary_Earned)
    # Salary_Earning_Details.objects.filter(
    #     month=Month,
    #     year=Year,
    #     OrganizationID=OID,
    #     EmpID=EmpID,
    #     IsDelete=False,
    #     SalaryTitle="Bonus"
    # ).update(Amount=Calculate_Bonus_Value)
    print("(CSSD Step 4.7.1")
    
    # Calculated_Employee_PF = Calculate_EmployeePF(Basic_Salary_Earned, fixed_basic)
    # Calculated_Employer_PF = Calculate_Employe_r_PF(total_earning, fixed_basic)
    # Calculated_ESIC = Calculate_ESIC(OID, fixed_gross, total_earning)
    # Calculated_ESIC = Calculate_ESIC(OID, fixed_gross, Data_PF_Calculation)
    Calculated_ESIC = Calculate_ESIC(OID, Basic_Salary_Earned)
    print("(CSSD Step 4.8")
    
    ESIC = Calculated_ESIC['ESIC']
    print("(CSSD Step 4.9")
    
    CompanyContributionToESIC = Calculated_ESIC['CompanyContributionToESIC']
    print("(CSSD Step 5")

    # === Overwrite Deduction Values ===
    attendance = SalaryAttendance.objects.filter(
        EmpID=EmpID, OrganizationID=OID, year=Year, month=Month, IsDelete=False
    ).first()

    if attendance:
        update_items = {
            "PT": Calculated_PT,
            "Employee PF @12% (Basic)": Calculated_Employee_PF,
            "ESIC @ 0.75%": ESIC,
        }

        for title, value in update_items.items():
            Salary_Deduction_Details.objects.filter(
                SalaryAttendance=attendance, SalaryTitle=title
            ).update(Amount=value)

        # Update Employer contribution in fixed
        Salary_Company_Contribution_Details.objects.filter(
            SalaryAttendance=attendance, SalaryTitle="Employer PF"
        ).update(Amount=Calculated_Employer_PF)

        Salary_Company_Contribution_Details.objects.filter(
            SalaryAttendance=attendance, SalaryTitle="Company Contribution to ESIC @ 3.25%"
        ).update(Amount=CompanyContributionToESIC)
        
    # Salary_Earning_Details.objects.filter(
    #     month=Month,
    #     year=Year,
    #     OrganizationID=OID,
    #     EmpID=EmpID,
    #     IsDelete=False,
    #     SalaryTitle="Bonus"
    # ).update(Amount=Calculate_Bonus_Value)


    # === Recalculate Total Deductions ===
    deduction_details = list(
        Salary_Deduction_Details.objects
        .filter(
            SalaryAttendance__EmpID=EmpID,
            SalaryAttendance__OrganizationID=OID,
            SalaryAttendance__year=Year,
            SalaryAttendance__month=Month,
            IsDelete=False
        )
        .values('SalaryTitle')
        .annotate(total=Sum('Amount'))
        .order_by('TitleOrder')
    )

    deduction_details = deduction_details[:-1]
    Deductions = {item['SalaryTitle']: float(item['total']) for item in deduction_details}

    print("(CSSD Step 6")

    # === Recalculate Total Company_Contribution ===
    Company_Contribution = list(
        Salary_Company_Contribution_Details.objects
        .filter(
            SalaryAttendance__EmpID=EmpID,
            SalaryAttendance__OrganizationID=OID,
            SalaryAttendance__year=Year,
            SalaryAttendance__month=Month,
            IsDelete=False
        )
        .values('SalaryTitle')
        .annotate(total=Sum('Amount'))
        .order_by('TitleOrder')
    )

    Company_Contribution = Company_Contribution[:-1]
    CompanyContribution = {item['SalaryTitle']: float(item['total']) for item in Company_Contribution}
    print("(CSSD Step 7")

    # total_Deductions = Decimal(math.ceil(sum(Deductions.values())))
    # total_CompanyContribution = Decimal(math.ceil(sum(CompanyContribution.values())))
    # net_payable = Decimal(math.ceil(float(total_earning - total_Deductions)))
    

    total_Deductions = Decimal(str(sum(Deductions.values()))).quantize(Decimal("0.00"))
    total_CompanyContribution = Decimal(str(sum(CompanyContribution.values()))).quantize(Decimal("0.00"))
    net_payable = (total_earning - total_Deductions).quantize(Decimal("0.00"))

    print("========----------------====================")
    print("Deductions::", Deductions)
    print("total_Deductions::", total_Deductions)
    
    for item in deduction_details:
        title = item['SalaryTitle']
        total = math.ceil(float(item['total']))

        print(F"{title}, ---, {total}")   # or use it anywhere

    print("========----------------====================")
    
    print("(CSSD Step 8")

    net_pay_int = int(net_payable)
    net_pay_in_words = num2words(net_pay_int, lang='en_IN').replace(',', '').title()
    net_pay_in_words = f"{net_pay_in_words} Rupees Only"
    CTC_Cal = total_CompanyContribution + total_earning
    # total_earning_int = float(total_earning, 2)
    # total_earning_float = float(f"{total_earning:.2f}")

    # net_payable = total_earning - total_Deductions
    print("(CSSD Step 9")

    slip = Salary_Slip_V1.objects.filter(
        EmpID=EmpID,
        OrganizationID=OID,
        year=Year,
        month=Month,
        IsDelete=False
    ).first()

    print("------------------------------------")
    # print("Total Earning is here:-- ->", total_earning)
    # print("total_earning_float is here:-- ->", total_earning_float)
    # print("Total Earning is here integer:-- ->", int(total_earning))
    # print("Total Company Contribution is here:-- ->", total_CompanyContribution)
    # print("CTC Calculation is here:-- ->", CTC_Cal)
    print("DEBUG total_earning:", total_earning, type(total_earning))
    print("DEBUG total_earning:", total_earning, type(total_earning))

    print("------------------------------------")
    if not isinstance(total_earning, Decimal):
        total_earning = Decimal(str(total_earning))
    print("(CSSD Step 10")

    if slip:
        slip.Total_Fixed = total_Fixed
        # slip.Total_Earning = total_earning.quantize(Decimal("0.00"))
        slip.Total_Earning = total_earning.quantize(Decimal("0.00"))
        # slip.Total_Earning = total_earning_float
        slip.Total_Deduction = total_Deductions
        slip.Total_Company_Contribution = total_CompanyContribution
        slip.CTC = CTC_Cal
        slip.Net_Payable = net_payable
        slip.Net_Payable_In_Words = net_pay_in_words
        slip.save(update_fields=["Total_Fixed", "Total_Earning", "Total_Deduction", "Total_Company_Contribution", "CTC", "Net_Payable", "Net_Payable_In_Words"])
    
    print("(CSSD Step 11 (Final)")

    # print(f"✅ Slip updated — Fixed: {total_Fixed}, Earning: {total_earning}, Deduction: {total_Deductions}, Net Pay: {net_payable}")



from django.shortcuts import redirect, render
from django.views.decorators.csrf import csrf_exempt

def safe_decimal(value):
    try:
        return Decimal(value)
    except (InvalidOperation, TypeError, ValueError):
        return Decimal('0.00')

def to_decimal(value):
    try:
        return Decimal(str(value))
    except:
        return Decimal(0)

TWOPLACES = Decimal("0.00")

def money(val):
    return Decimal(str(val)).quantize(TWOPLACES, rounding=ROUND_HALF_UP)

@csrf_exempt  # remove if you're using {% csrf_token %}
def update_salary_slip_manual_fields(request):
    if request.method == "POST":
        slip_id = request.POST.get('slip_id')

        arrear = safe_decimal(request.POST.get('arrear'))
        advance_loan = safe_decimal(request.POST.get('advance_loan'))
        tax_deduction = safe_decimal(request.POST.get('tax_deduction'))
        other_deduction = safe_decimal(request.POST.get('other_deduction'))
        print("arrears:", arrear)
        print("advance_loan:", advance_loan)
        print("tax_deduction:", tax_deduction)
        print("other_deduction:", other_deduction)

        # Example: Update the salary slip record
        slip = Salary_Slip_V1.objects.filter(id=slip_id, IsDelete=False).first()
        if not slip:
            return redirect('Show_Salary_Slip_PDF', EmpID=slip.EmpID, OID=slip.OrganizationID, Year=slip.year, Month=slip.month)
        
        EmpID = slip.EmpID
        OID = slip.OrganizationID
        Year = slip.year
        Month = slip.month

        # === Earnings ===
        earning_details = list(
            Salary_Earning_Details.objects
            .filter(
                month=Month,
                year=Year,
                OrganizationID=OID,
                EmpID=EmpID,
                IsDelete=False
            )
            .values('SalaryTitle', 'TitleOrder')
            .annotate(total=Sum('Amount'))
            .order_by('TitleOrder')
        )
        earning_details = earning_details[:-1]
        # Earning = {item['SalaryTitle']: float(item['total']) for item in earning_details}
        Earning = {
            item['SalaryTitle']: Decimal(str(item['total']))
            for item in earning_details
        }

        # === Totals ===
        # Total_Earning = round_decimal(sum(Earning.values()))
        Total_Earning = sum(Earning.values(), Decimal("0.00"))
        print("Total_Earning:",Total_Earning)
        


        # === Deductions ===
        deduction_details = list(
            Salary_Deduction_Details.objects.filter(
                SalaryAttendance__EmpID=EmpID,
                SalaryAttendance__OrganizationID=OID,
                SalaryAttendance__year=Year,
                SalaryAttendance__month=Month,
                IsDelete=False
            )
            .values('SalaryTitle')
            .annotate(total=Sum('Amount'))
            .order_by('TitleOrder')
        )
        deduction_details = deduction_details[:-1]
        # Deductions = {item['SalaryTitle']: float(item['total']) for item in deduction_details}
        Deductions = {
            item['SalaryTitle']: Decimal(str(item['total']))
            for item in deduction_details
        }

        # Total_Deduction = Decimal(sum(Deductions.values()))
        Total_Deduction = sum(Deductions.values(), Decimal("0.00"))
        print("total deduction:",Total_Deduction)
        

        Total_Earning_revised = Total_Earning + to_decimal(arrear)
        print("Total_Earning_revised:",Total_Earning_revised)
        

        slip_Total_Deduction = to_decimal(slip.Total_Deduction)
        slip_Total_Company_Contribution = to_decimal(slip.Total_Company_Contribution)

        Total_Deduction_revised = to_decimal(Total_Deduction) + to_decimal(advance_loan) + to_decimal(tax_deduction) + to_decimal(other_deduction)
        print(f"Total_Deduction_revised:{Total_Deduction_revised} = {Total_Deduction} + {advance_loan} + {tax_deduction} + {other_deduction}")
        # print("Total_Deduction_revised:",Total_Deduction_revised)
        

        Net_Payable_revised = Total_Earning_revised - Total_Deduction_revised
        print(f"Net_Payable_revised:{Net_Payable_revised} = {Total_Earning_revised} - {Total_Deduction_revised}")
        
        # print("-----------------------------------------------------")
        # print("other_deduction::",other_deduction)
        # print("slip_Total_Deduction::",slip_Total_Deduction)
        # print("Total_Deduction_revised::",Total_Deduction_revised)
        # print("Net_Payable_revised::",Net_Payable_revised)
        # print("-----------------------------------------------------")

        CTC_Cal = Total_Earning_revised + slip_Total_Company_Contribution
        print(f"CTC_Cal:{CTC_Cal} = {Total_Earning_revised} + {slip_Total_Company_Contribution}")
        


        net_pay_int = int(Net_Payable_revised)
        net_pay_in_words = num2words(net_pay_int, lang='en_IN').replace(',', '').title()
        net_pay_in_words = f"{net_pay_in_words} Rupees Only"

        # revised fields
        arrear = money(arrear)
        advance_loan = money(advance_loan)
        tax_deduction = money(tax_deduction)
        other_deduction = money(other_deduction)

        Total_Earning = money(Total_Earning)
        Total_Deduction = money(Total_Deduction)

        Total_Earning_revised = money(Total_Earning + arrear)
        Total_Deduction_revised = money(
            Total_Deduction + advance_loan + tax_deduction + other_deduction
        )

        Net_Payable_revised = money(Total_Earning_revised - Total_Deduction_revised)

        CTC_Cal = money(Total_Earning_revised + slip_Total_Company_Contribution)

        if slip:
            slip.Arrers = arrear
            slip.Advance_Loan = advance_loan
            slip.Tax_Deduction = tax_deduction
            slip.Other_Deduction = other_deduction

            slip.Total_Earning = Total_Earning_revised
            slip.Total_Deduction = Total_Deduction_revised
            slip.CTC = CTC_Cal
            slip.Net_Payable = Net_Payable_revised
            slip.Net_Payable_In_Words = net_pay_in_words
            try:
                slip.save(update_fields=["Arrers", "Advance_Loan", "Tax_Deduction", "Other_Deduction", "Total_Earning", "Total_Deduction", "Net_Payable", "Net_Payable_In_Words", "CTC"])
            except Exception as e:
                # print(f"Error saving slip: {e}")
                return JsonResponse({
                    "status": "error",
                    "message": f"Failed to update salary slip, {{str(e)}}"
                }, status=500)
        else:
            return JsonResponse({
                "status": "error",
                "message": "Salary slip not found"
            }, status=404)
        
    return JsonResponse({
        "status": "success",
        "message": "Salary computed with split periods"
    }, status=200)




# Delete_Per_Day_Salary
def Delete_Per_Day_Salary(EmpID, OID, Year, Month):
    """
        Completely removes all forward salary-related entries for the given employee
        for the specified month/year. Keeps data integrity clean before re-generation.
    """
    
    # === Recalculate Total Deductions ===
    if EmpID is None or OID is None or Year is None or Month is None:
        # print("Missing parameters for deletion")
        return
    
    Salary_Earning_Details.objects.filter(
            # SalaryAttendance__EmpID=EmpID,
            # SalaryAttendance__OrganizationID=OID,
            # SalaryAttendance__year=Year,
            # SalaryAttendance__month=Month,
            # IsDelete=False
            month=Month,
            year=Year,
            OrganizationID=OID,
            EmpID=EmpID,
            IsDelete=False
        ).delete()
    
    Salary_Fixed_Details.objects.filter(
            SalaryAttendance__EmpID=EmpID,
            SalaryAttendance__OrganizationID=OID,
            SalaryAttendance__year=Year,
            SalaryAttendance__month=Month,
            IsDelete=False
        ).delete()
    
    Salary_Deduction_Details.objects.filter(
            SalaryAttendance__EmpID=EmpID,
            SalaryAttendance__OrganizationID=OID,
            SalaryAttendance__year=Year,
            SalaryAttendance__month=Month,
            IsDelete=False
        ).delete()
    
    Salary_Company_Contribution_Details.objects.filter(
            SalaryAttendance__EmpID=EmpID,
            SalaryAttendance__OrganizationID=OID,
            SalaryAttendance__year=Year,
            SalaryAttendance__month=Month,
            IsDelete=False
        ).delete()
    
    Response_Data_Generate_Slip.objects.filter(
        Month_No=Month,
        Year=Year,
        Emp_ID=EmpID,
        Hotel_Id=OID
    ).delete()

    # return JsonResponse({
    #     "status": "success",
    #     "message": "Deleted per day salary components"
    # }, status=200)
    # print("All per-day salary, fixed, and deduction records deleted.")
    return True



def Response_Data_Slip_Api(request):
    current_date = datetime.now()

    year = int(request.GET.get('year', current_date.year))
    month_no = int(request.GET.get('month_no', current_date.month))
    # EmpCode = request.GET.get('EmpCode')
    OID = request.GET.get('OID')

    data = Response_Data_Generate_Slip.objects.filter(
        IsDelete=False, Hotel_Id=OID, Month_No=month_no, Year=year
    ).values("Emp_Name", "Emp_Code", "Response_Message", "Response_Status")

    return JsonResponse({
        "success": True,
        "count": data.count(),
        "Response_Data": list(data),
    }, status=200)



from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import calendar
from decimal import Decimal


import openpyxl
from django.http import HttpResponse
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

def Export_Salary_Slip_Excel(request, EmpID, OID, Year, Month):
    # === Fetch Basic Slip Information ===
    slip = Salary_Slip_V1.objects.filter(
        EmpID=EmpID,
        OrganizationID=OID,
        year=Year,
        month=Month,
        IsDelete=False
    ).first()

    if not slip:
        return HttpResponse("No salary slip found for this employee.", status=404)
    
    fixed_details = list(
        Salary_Fixed_Details.objects
        .filter(
            SalaryAttendance__EmpID=EmpID,
            SalaryAttendance__OrganizationID=OID,
            SalaryAttendance__year=Year,
            SalaryAttendance__month=Month,
            IsDelete=False
        )
        .values('SalaryTitle', 'TitleOrder')
        .annotate(total=Sum('Amount'))
        .order_by('TitleOrder')
    )
    fixed_details = fixed_details[:-1]
    Fixed = {item['SalaryTitle']: float(item['total']) for item in fixed_details}
    
    earning_details = list(
        Salary_Earning_Details.objects
        .filter(
            SalaryAttendance__EmpID=EmpID,
            SalaryAttendance__OrganizationID=OID,
            SalaryAttendance__year=Year,
            SalaryAttendance__month=Month,
            IsDelete=False
        )
        .values('SalaryTitle', 'TitleOrder')
        .annotate(total=Sum('Amount'))
        .order_by('TitleOrder')
    )
    earning_details = fixed_details[:-1]
    # Fixed = {item['SalaryTitle']: float(item['total']) for item in fixed_details}


    deduction_details = list(
        Salary_Deduction_Details.objects
        .filter(
            SalaryAttendance__EmpID=EmpID,
            SalaryAttendance__OrganizationID=OID,
            SalaryAttendance__year=Year,
            SalaryAttendance__month=Month,
            IsDelete=False
        )
        .values('SalaryTitle')
        .annotate(total=Sum('Amount'))
        .order_by('TitleOrder')
    )

    deduction_details = deduction_details[:-1]
    # Deductions = {item['SalaryTitle']: float(item['total']) for item in deduction_details}

    # Convert to dictionaries
    Fixed = {item['SalaryTitle']: round(float(item['total']), 2) for item in fixed_details}
    Earning = {item['SalaryTitle']: round(float(item['total']), 2) for item in earning_details}
    Deductions = {item['SalaryTitle']: round(float(item['total']), 2) for item in deduction_details}

    # === Combine Data ===
    all_titles = list(Fixed.keys()) + list(Earning.keys()) + list(Deductions.keys())

    # === Create Workbook ===
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Salary Slip"

    # === Header Row ===
    headers = ["Emp Name", "Emp Code", "Designation", "Department", "EmpStatus", 'Advance_Loan', 'Tax_Deduction', 'Other_Deduction', 'Other_Deduction', 'Total_Deduction', 'Net_Payable', 'Net_Payable_In_Words', 'Total_No_Days_In_Month', 'Present_Days'] + all_titles
    ws.append(headers)

    # === Data Row ===
    data_row = [slip.Emp_Name, slip.EmployeeCode, slip.Designation, slip.Department, slip.EmpStatus, slip.Advance_Loan, slip.Tax_Deduction, slip.Other_Deduction, slip.Other_Deduction, slip.Total_Deduction, slip.Net_Payable, slip.Net_Payable_In_Words, slip.Total_No_Days_In_Month, slip.Present_Days]
    for title in all_titles:
        amount = Earning.get(title) or Deductions.get(title) or Fixed.get(title) or 0
        data_row.append(amount)
    ws.append(data_row)

    # === Styling ===
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = 18

    # === Prepare Response ===
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"SalarySlip_{slip.EmployeeCode}_{Year}_{Month}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response





# def Verify_Salary_Slip(request, EmpID, OID, Year, Month):
def Verify_Salary_Slip(request):
    current_date = datetime.now()
    OrganizationID = request.session["OrganizationID"]
    Year = int(request.GET.get('year', current_date.year))
    Month = int(request.GET.get('month_no', current_date.month))
    today = datetime.today()
    CYear = today.year
    CMonth = today.month
    context = {
        'OrganizationID':OrganizationID,
        'Year':Year,
        'CYear':range(CYear,2020,-1),
        'CMonth':CMonth,
        'Month':Month
    }
    return render(request, "EMP_PAY/MoveToPayroll_Template/Verify_Salary_Slip_Data.html", context)



from rest_framework.decorators import api_view

@api_view(['GET'])
def Verify_Salary_Slip_API(request):
    try:
        OID = request.GET.get('OID')
        Year = request.GET.get('Year')
        Month = request.GET.get('Month')
        Status = request.GET.get('Status')
        UserDepartment = (request.session.get("Department_Name") or "").lower()

        # print("my status is here", Status)
        # print("Type of status is here", type(Status))


        if not all([OID, Year, Month]):
            return Response({"error": "OID, Year, and Month are required."},
                            status=status.HTTP_400_BAD_REQUEST)

        Year, Month = int(Year), int(Month)
        # ---- Base filter ----
        filters = {
            "OrganizationID": OID,
            "year": Year,
            "month": Month,
            "IsDelete": False,
        }

        lastStatus = 0

        # ---- Apply status logic ----
        if Status != '2':
            verify_field = None
            if UserDepartment == 'hr':
                verify_field = 'HrVerify'
                lastStatus=Status
            elif UserDepartment == 'finance':
                verify_field = 'FcVerify'
                lastStatus=Status

            if verify_field:
                slips = Salary_Slip_V1.objects.filter(
                    OrganizationID=OID,
                    year=Year,
                    month=Month,
                    IsDelete=False,
                    **{verify_field: Status == '1'}
                )
            else:
                print("dept is other", UserDepartment)
        else:
            slips = Salary_Slip_V1.objects.filter(
                OrganizationID=OID,
                year=Year,
                month=Month,
                IsDelete=False
            )

        # slips = Salary_Slip_V1.objects.filter(**filters)


        if not slips.exists():
            return Response({"message": "No salary slips found."},
                            status=status.HTTP_404_NOT_FOUND)

        results = []

        for slip in slips:
            slip_data = {
                "SlipID": slip.id,
                "month": slip.month,
                "year": slip.year,
                "EmployeeCode": slip.EmployeeCode,
                "Emp_Name": slip.Emp_Name,
                "Designation": slip.Designation,
                "Department": slip.Department,
                "EmpStatus": slip.EmpStatus,
                "DOJ": slip.DOJ,
                "ESIC_Number": slip.ESIC_Number,
                "Provident_Fund_Number": str(slip.Provident_Fund_Number),

                "EmpID": slip.EmpID,
                "OrganizationID": slip.OrganizationID,
                "Year": slip.year,
                "Month": slip.month,


                "Total_Fixed": slip.Total_Fixed,
                "Total_Earning": slip.Total_Earning,
                "Total_Deduction": slip.Total_Deduction,
                "Total_Company_Contribution": slip.Total_Company_Contribution,
                "NetSalary": slip.Net_Payable,
                "CTC": slip.CTC,

                "Arrers": slip.Arrers,
                "Advance_Loan": slip.Advance_Loan,
                "Tax_Deduction": slip.Tax_Deduction,
                "Other_Deduction": slip.Other_Deduction,
                "Total_Deduction": slip.Total_Deduction,
                "Present_Days": slip.Present_Days,
                "Absent_Days": slip.Absent_Days,
                "Total_No_Days_In_Month": slip.Total_No_Days_In_Month,

                "Bank_Name": slip.Bank_Name,
                "Bank_Account_Number": str(slip.Bank_Account_Number),
                "Bank_IFSC_Code": slip.Bank_IFSC_Code,
                "Bank_Branch": slip.Bank_Branch,

                "HrVerify": slip.HrVerify,
                "FcVerify": slip.FcVerify,
                "IsLocked": slip.IsLocked,
                "IsGenerated": slip.IsGenerated,
                "lastStatus": lastStatus,
            }

            # --- Fetch breakdown per employee ---
            Fixed = {
                item['SalaryTitle']: math.ceil(float(item['total']))
                for item in Salary_Fixed_Details.objects
                    .filter(SalaryAttendance__EmpID=slip.EmpID,
                            SalaryAttendance__OrganizationID=OID,
                            SalaryAttendance__year=Year,
                            SalaryAttendance__month=Month,
                            IsDelete=False)
                    .values('SalaryTitle')
                    .annotate(total=Sum('Amount'))
                    .order_by('TitleOrder')
            }

            Earning = {
                # item['SalaryTitle']: int(item['total'])
                item['SalaryTitle']: float(f"{item['total']:.2f}")
                for item in Salary_Earning_Details.objects
                    .filter(
                        # SalaryAttendance__EmpID=slip.EmpID,
                        #     SalaryAttendance__OrganizationID=OID,
                        #     SalaryAttendance__year=Year,
                        #     SalaryAttendance__month=Month,
                        #     IsDelete=False
                        
                        month=Month,
                        year=Year,
                        OrganizationID=OID,
                        EmpID=slip.EmpID,
                        IsDelete=False
                    )
                    .values('SalaryTitle')
                    .annotate(total=Sum('Amount'))
                    .order_by('TitleOrder')
            }

            Deductions = {
                item['SalaryTitle']: math.ceil(float(item['total']))
                for item in Salary_Deduction_Details.objects
                    .filter(SalaryAttendance__EmpID=slip.EmpID,
                            SalaryAttendance__OrganizationID=OID,
                            SalaryAttendance__year=Year,
                            SalaryAttendance__month=Month,
                            IsDelete=False)
                    .values('SalaryTitle')
                    .annotate(total=Sum('Amount'))
                    .order_by('TitleOrder')
            }

            Company_Contribution = {
                item['SalaryTitle']: math.ceil(float(item['total']))
                for item in Salary_Company_Contribution_Details.objects
                    .filter(SalaryAttendance__EmpID=slip.EmpID,
                            SalaryAttendance__OrganizationID=OID,
                            SalaryAttendance__year=Year,
                            SalaryAttendance__month=Month,
                            IsDelete=False)
                    .values('SalaryTitle')
                    .annotate(total=Sum('Amount'))
                    .order_by('TitleOrder')
            }

            # con_qs = Salary_Company_Contribution_Details.objects.filter(
            #     SalaryAttendance__EmpID = slip.EmpID,
            #     SalaryAttendance__OrganizationID = OID,
            #     SalaryAttendance__year = Year,
            #     SalaryAttendance__month = Month,
            #     IsDelete=False
            # )

            # print("Raw CC Queryset Count:", Company_Contribution.count())
            # print("Raw CC Queryset Count:", Company_Contribution)
            # print("Raw CC Rows:", list(con_qs.values()))


            # print("<------------------------                 -------------------------->")
            # print("Companyt contribution is here    ::--->", Company_Contribution)
            # print("EmployeeCode is here::--->", slip.EmployeeCode)
            # print("EmpID is here::--->", slip.EmpID)
            # print("OrganiZationID is here::--->", OID)
            # print("Year is here::--->", Year)
            # print("Month is here::--->", Month)
            # print("<------------------------                 -------------------------->")

            # total_fixed = sum(Fixed.values())
            # total_earning = sum(Earning.values())
            # total_deduction = sum(Deductions.values())
            # net_pay = Decimal(total_earning - total_deduction).quantize(Decimal('0.01'))

            results.append({
                "Slip": slip_data,
                "Fixed": Fixed,
                "Earning": Earning,
                "Deductions": Deductions,
                "Company_Contribution": Company_Contribution,
                # "con_qs": con_qs,
                # "TotalFixed": total_fixed,
                # "TotalEarning": total_earning,
                # "TotalDeduction": total_deduction,
                # "NetPay": float(net_pay)
            })

        return Response({
            "OrganizationID": OID,
            "Year": Year,
            "Month": Month,
            "MonthName": calendar.month_name[Month],
            "Employees": results
        }, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)




def Bulk_Verify_Salary_Slip(request):
    if request.method == "POST":
        UserID = str(request.session.get("UserID", 0))
        OrganizationID = request.session.get("OrganizationID", 0)

        employees_json = request.POST.get("employees")
        if not employees_json:
            return JsonResponse({"error": "No employees provided"}, status=400)

        try:
            employees = json.loads(employees_json)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON format"}, status=400)

        results = []
        updated_count = 0

        for emp in employees:
            year = emp.get("year")
            month_no = emp.get("month_no")
            SlipID = emp.get("SlipID")
            PressedButton = emp.get("PressedButton")
            lastStatus = emp.get("lastStatus") or 0

            if not SlipID:
                continue

            slips = Salary_Slip_V1.objects.filter(
                id=SlipID,
                year=year,
                month=month_no,
                IsDelete=False
            )

            if not slips.exists():
                results.append({
                    "SlipID": SlipID,
                    "status": "not_found"
                })
                continue

            if PressedButton == 'HRBulkVerify':
                slips.update(HrVerify=True)
                status_text = "HR Verified"
            elif PressedButton == 'FCBulkVerify':
                slips.update(FcVerify=True)
                status_text = "FC Verified"
            else:
                status_text = "Unknown Action"
            
            # slips.save()

            updated_count += slips.count()
            results.append({
                "SlipID": SlipID,
                "status": status_text
            })

        return JsonResponse({
            "year": year,
            "lastStatus": lastStatus,
            "month_no": month_no,
            "OrganizationID": OrganizationID,
            "updated_count": updated_count,
            "results": results,
            "success": True,
            "message": f"Processed {len(employees)} employees successfully."
        })





def Audit_Attendance_Report(request):
    if 'OrganizationID' not in request.session:
        return redirect(MasterAttribute.Host)
    else:
        print("Show Page Session")

    OrganizationID = request.session["OrganizationID"]
      
    current_date = datetime.now()
    year = int(request.GET.get('year', current_date.year))
    month_no = int(request.GET.get('month_no', current_date.month))


    start_date = date(year, month_no, 1)
    last_day = calendar.monthrange(year, month_no)[1]  # returns (weekday, num_days)
    end_date = date(year, month_no, last_day)

    Org_Name = OrganizationMaster.objects.filter(
        OrganizationID=OrganizationID, IsDelete=False, IsNileHotel=True
    ).only('OrganizationName').first()

    today = datetime.today()
    CYear = today.year
    CMonth = today.month

    # Get employee list
    emp_list = EmployeeDataSelectForSalary(OrganizationID, month_no, year)

    # Get all salary attendance records for the month
    salary_attendance_qs = SalaryAttendance.objects.filter(
        month=month_no,
        year=year,
        OrganizationID=OrganizationID,
        IsDelete=False,
        IsAttendanceMoved=True,
        IsAttendanceModified=True
    ).values('EmployeeCode', 'Status', 'ActualStatus', 'ModifyDateTime', 'Date', 'Remarks')

    # Group all attendance records per EmployeeCode
    salary_attendance_map = {}
    for item in salary_attendance_qs:
        emp_code = item['EmployeeCode']
        if emp_code not in salary_attendance_map:
            salary_attendance_map[emp_code] = []
        salary_attendance_map[emp_code].append({
            'Status': item['Status'],
            'Date': item['Date'],
            'ModifyDateTime': item['ModifyDateTime'],
            'Remarks': item['Remarks'],
            'ActualStatus': item['ActualStatus']
        })

    # Build final employee list with all their attendance records
    final_emp_list = []
    for emp in emp_list:
        emp_code = emp['EmployeeCode']
        if emp_code in salary_attendance_map:
            final_emp_list.append({
                'EmployeeCode': emp['EmployeeCode'],
                'EmpName': emp['EmpName'],
                'attendance_records': salary_attendance_map[emp_code]  # all records for this employee
            })

    # Month navigation
    previous_month = (datetime(year, month_no, 1) - timedelta(days=1)).strftime('%Y-%m-%d')
    next_month_date = min(
        datetime(year, month_no, 28) + timedelta(days=4),
        datetime.now().replace(day=1)
    )
    if next_month_date.month == month_no:
        next_month_date = next_month_date.replace(day=1)
    next_month = next_month_date.strftime('%Y-%m-%d')


    context = {
        'Session_OrganizationID':OrganizationID,
        'emps': final_emp_list,
        'current_month': datetime(year, month_no, 1).strftime('%Y-%m-%d'),
        'previous_month': previous_month,
        'next_month': next_month,
        'month_no': month_no,
        'year': year,
        'Org_Name': Org_Name,
        'today': today,
        'StartDate': start_date,
        'EndDate': end_date,
        'CMonth': month_no,
        'CYear':range(CYear,2020,-1),
    }
    return render(request, "EMP_PAY/Reports_Page/Audit_Attendance_Report.html", context)








def Excel_Attendance_Upload_View(request):
    if 'OrganizationID' not in request.session:
        return redirect(MasterAttribute.Host)
    
    context = {}
    # return render(request, "EMP_PAY/Attendance/Alifupload_csv.html", {'form':form})
    return render(request, "EMP_PAY/MoveToPayroll_Template/Excel_Attendance_Upload.html", context)




import pandas as pd
from datetime import datetime
from django.shortcuts import render
from django.contrib import messages
from django.db import transaction
from .models import SalaryAttendance


# def Excel_Attendance_Upload_CSV_Api(request):
#     if request.method != "POST":
#         print("request method is not post")
#         return JsonResponse({"status": "error", "message": "Invalid request"}, status=400)
    

#     print("request method is post")
#     excel_file = request.FILES.get("excel_file")
#     print("FILES:", request.FILES)

#     if not excel_file:
#         # return JsonResponse({
#         #     "status": "error",
#         #     "message": "No file uploaded"
#         # }, status=400)
        
#         return JsonResponse({
#             "status": "error",
#             "message": str(e),
#             "errors": []
#         }, status=500)


#     try:
#         # df = pd.read_excel(excel_file)
#         df = pd.read_excel(excel_file, header=6)
#         df.columns = df.columns.str.strip().str.upper()

        
#         print("df are there::", df)
#         print("recived columns are there::", df.columns)

#         required_columns = {"ID", "INDATE", "INTIME", "OUTTIME", "DAYTOTAL", "DAYSTATUS"}
#         if not required_columns.issubset(df.columns):
#             return JsonResponse({
#                 "status": "error",
#                 "message": f"Missing columns: {required_columns - set(df.columns)}"
#             }, status=400)

#         updated_count = 0
#         errors = []

#         with transaction.atomic():
#             for index, row in df.iterrows():
#                 try:
#                     employee_code = str(row["ID"]).strip()
#                     # attendance_date = pd.to_datetime(row["INDATE"]).date()
#                     attendance_date = pd.to_datetime(
#                         row["INDATE"],
#                         dayfirst=True,
#                         errors="coerce"
#                     ).date()
#                     print("attendence date:", attendance_date)

#                     attendance = Excel_Attendance_Upload_Punch_Record.objects.filter(
#                         EmployeeCode=employee_code,
#                         Date=attendance_date,
#                         IsDelete=False
#                     ).first()

#                     if not attendance:
#                         errors.append(f"Row {index+2}: Record not found")
#                         continue

#                     attendance.In_Time = str(row["INTIME"]) if pd.notna(row["INTIME"]) else None
#                     attendance.Out_Time = str(row["OUTTIME"]) if pd.notna(row["OUTTIME"]) else None
#                     attendance.Duty_Hour = str(row["DAYTOTAL"]) if pd.notna(row["DAYTOTAL"]) else None
#                     attendance.Status = str(row["DAYSTATUS"]) if pd.notna(row["DAYSTATUS"]) else None

#                     attendance.IsUpload = True
#                     attendance.IsAttendanceModified = True
#                     attendance.save()
#                     print("the data is saved")

#                     updated_count += 1

#                 except Exception as row_err:
#                     print("the data is not saved")
#                     errors.append(f"Row {index+2}: {str(row_err)}")

#         return JsonResponse({
#             "status": "success",
#             "updated_records": updated_count,
#             "errors": errors
#         })

#     except Exception as e:
#         return JsonResponse({
#             "status": "error",
#             "message": str(e)
#         }, status=500)


from django.http import JsonResponse

def Excel_Attendance_Upload_CSV_Api(request):
    if request.method != "POST":
        # print("request method is not post")
        return JsonResponse({"status": "error", "message": "Invalid request"}, status=400)
    
    OID = request.session.get("OrganizationID") or 0
    

    excel_file = request.FILES.get("excel_file")

    if not excel_file:
        return JsonResponse({
            "status": "error",
            "message": str(e),
            "errors": []
        }, status=500)


    try:
        # df = pd.read_excel(excel_file)
        df = pd.read_excel(excel_file, header=6)
        df.columns = df.columns.str.strip().str.upper()

        required_columns = {"ID", "INDATE", "INTIME", "OUTTIME", "DAYTOTAL", "DAYSTATUS"}
        if not required_columns.issubset(df.columns):
            return JsonResponse({
                "status": "error",
                "message": f"Missing columns: {required_columns - set(df.columns)}"
            }, status=400)

        created_count = 0
        updated_count = 0
        errors = []

        with transaction.atomic():
            for index, row in df.iterrows():
                try:
                    employee_code = str(row["ID"]).strip()
                    # attendance_date = pd.to_datetime(row["INDATE"]).date()
                    attendance_date = pd.to_datetime(
                        row["INDATE"],
                        dayfirst=True,
                        errors="coerce"
                    ).date()

                    attendance, created = Excel_Attendance_Upload_Punch_Record.objects.get_or_create(
                        EmployeeCode=employee_code,
                        Date=attendance_date,
                        OrganizationID=OID,
                        defaults={
                            "IsDelete": False,
                        }
                    )
                    
                    STATUS_MAP = {
                        "A": "Absent",
                        "P": "Present",
                    }
                    
                    attendance.In_Time = str(row["INTIME"]) if pd.notna(row["INTIME"]) else None
                    attendance.Out_Time = str(row["OUTTIME"]) if pd.notna(row["OUTTIME"]) else None
                    attendance.Duty_Hour = str(row["DAYTOTAL"]) if pd.notna(row["DAYTOTAL"]) else None
                    # attendance.Status = str(row["DAYSTATUS"]) if pd.notna(row["DAYSTATUS"]) else None
                    
                    raw_status = str(row["DAYSTATUS"]).strip().upper() if pd.notna(row["DAYSTATUS"]) else None
                    attendance.Status = STATUS_MAP.get(raw_status, raw_status)


                    attendance.IsUpload = True
                    attendance.IsAttendanceModified = True
                    attendance.OrganizationID = OID
                    attendance.save()
                    

                    if created:
                        created_count += 1
                    else:
                        updated_count += 1
                        
                    # 🔥 Sync to Attendance_Data
                    sync_attendance_async(attendance)
                        
                except Exception as row_err:
                    print("the data is not saved")
                    errors.append(f"Row {index+2}: {str(row_err)}")

        return JsonResponse({
            "status": "success",
            "created_records": created_count,
            "updated_records": updated_count,
            "errors": errors
        })
    except Exception as e:
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=500)


from django.utils import timezone
from .models import Attendance_Data


import threading

def sync_attendance_async(attendance):
    thread = threading.Thread(
        target=sync_attendance_data_from_punch,
        args=(attendance,)
    )
    thread.daemon = True
    thread.start()


def sync_attendance_data_from_punch(attendance):
    """
    Sync Excel_Attendance_Upload_Punch_Record
    to Attendance_Data table
    """

    attendance_data, created = Attendance_Data.objects.get_or_create(
        EmployeeCode=attendance.EmployeeCode,
        Date=attendance.Date,
        defaults={
            "IsDelete": False,
            "OrganizationID": attendance.OrganizationID,
            "CreatedDateTime": timezone.now(),
        }
    )

    attendance_data.In_Time = attendance.In_Time
    attendance_data.Out_Time = attendance.Out_Time
    attendance_data.Duty_Hour = attendance.Duty_Hour
    attendance_data.Status = attendance.Status
    attendance_data.IsUpload = True
    attendance_data.ModifyDateTime = timezone.now()

    attendance_data.save()

    return attendance_data, created
